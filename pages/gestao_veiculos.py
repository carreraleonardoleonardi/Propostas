import streamlit as st
import pandas as pd
import requests
import json
import datetime
import io

# ── Constantes ───────────────────────────────────────────────────────────────

GV_SHEET_URL = "https://docs.google.com/spreadsheets/d/1BpAtiXz4AEuQg4kVx8OFonohPlvbScdOgWPIZRxQnxo/export?format=csv&gid=461042346"
GV_WEBHOOK   = "https://script.google.com/macros/s/AKfycbzFP-ezBsVx7W7VhYATKgaqdAg485o0AQb8s9FdGTlvmdzK1YRj7dCUVfTrXNgJOToc/exec"
SENHA_FECHAMENTO = "#FECHAMENTO"

GV_STATUS_LIST = [
    "Trânsito Livre", "Trânsito Vendido", "Livre",
    "Aguardando Atribuição", "Aguardando Agendamento", "Agendado",
    "Entregue", "Reagendar", "Avariado", "Distrato",
    "Remoção", "Reserva Temporária", "Evento Signature",
]
GV_FABRICANTES  =  ["Volkswagen","Chevrolet","Nissan","Jeep","GWM","GAC","Omoda", "Renault","Hyundai","Toyota","Fiat","Ford","Honda","Citroën","Peugeot","Mitsubishi","Subaru","Chery","JAC","Lexus","Kia","Dodge","BMW","Mercedes-Benz","Audi","Porsche","Volvo","Mini","Land Rover","Jaguar","Alfa Romeo","Aston Martin","Bentley","Rolls-Royce","McLaren","Pagani","Bugatti","Koenigsegg","Zeekr","BYD","Leapmotor"]
GV_LOCADORAS    = ["LM FROTAS","RCI","TOOT","GM Fleet", "Arval", "Localiza"]
GV_LOJAS        = ["LOJA ALPHAVILLE","LOJA VILLA LOBOS","LOJA OSASCO","LOJA BUTANTÃ","LOJA COTIA","OUTRO DN"]
GV_COMBUSTIVEIS = ["Flex","Gasolina","Elétrico","Híbrido","Diesel"]
GV_COLUNAS      = [
    "id","fabricante","modelo","chassi","placa","cor",
    "ano_fabricacao","ano_modelo","combustivel","opcionais",
    "locadora","consultor","cliente","pedido","status",
    "local_atual","data_chegada","data_entrega","hora_entrega",
    "entregador","avaria","obs_avaria","loja_entrega",
    "valor_nf","margem","comissao",
    "criado_em","atualizado_em","atualizado_por",
    "transporte_solicitado",
]

STATUS_CORES = {
    "Livre":                  "#22c55e",
    "Trânsito Livre":         "#3b82f6",
    "Trânsito Vendido":       "#8b5cf6",
    "Aguardando Atribuição":  "#eab308",
    "Aguardando Agendamento": "#f97316",
    "Agendado":               "#06b6d4",
    "Entregue":               "#10b981",
    "Reagendar":              "#f59e0b",
    "Avariado":               "#ef4444",
    "Distrato":               "#6b7280",
    "Remoção":                "#64748b",
    "Reserva Temporária":     "#a855f7",
    "Evento Signature":       "#ec4899",
}

# Paleta oficial Carrera
D_ESC    = "#b57b3f"
D_CLR    = "#dfc28a"
AZUL     = "#213144"

# ── Helpers ──────────────────────────────────────────────────────────────────

def parse_data(valor):
    if not valor or str(valor).strip() in ("","nan","None","NaT"): return None
    s = str(valor).strip()
    for fmt in ("%d/%m/%Y","%Y-%m-%d","%Y/%m/%d","%d-%m-%Y"):
        try: return datetime.datetime.strptime(s, fmt).date()
        except: continue
    return None

def fmt_data(valor):
    d = parse_data(valor)
    return d.strftime("%d/%m/%Y") if d else ""

def calcular_idade(row):
    chegada = parse_data(str(row.get("data_chegada","")))
    if not chegada: return None
    if str(row.get("status","")).strip() == "Entregue":
        entrega = parse_data(str(row.get("data_entrega","")))
        ref = entrega or datetime.date.today()
    else:
        ref = datetime.date.today()
    return (ref - chegada).days

def farol_idade(dias):
    if dias is None: return "⚪"
    if dias <= 20:   return "🟢"
    if dias <= 30:   return "🟡"
    if dias <= 45:   return "🔴"

def farol_agendamento(row):
    if str(row.get("status","")).strip() == "Entregue": return "🟢"
    data_ent = parse_data(str(row.get("data_entrega","")))
    if not data_ent: return "⚪"
    hora_str = str(row.get("hora_entrega","")).strip()
    agora = datetime.datetime.now()
    try:
        h, m = map(int, hora_str.split(":"))
        dt_ag = datetime.datetime.combine(data_ent, datetime.time(h, m))
    except:
        dt_ag = datetime.datetime.combine(data_ent, datetime.time(23, 59))
    return "🔴" if agora > dt_ag else "🟡"

RODIZIO_SP       = {"1":0,"2":0,"3":1,"4":1,"5":2,"6":2,"7":3,"8":3,"9":4,"0":4}
RODIZIO_BLOQUEIO = [(datetime.time(6,0), datetime.time(10,0)), (datetime.time(16,0), datetime.time(20,0))]

def verificar_rodizio(placa, data, hora):
    if not placa: return None
    final = placa.strip()[-1].upper()
    if final not in RODIZIO_SP or data.weekday() != RODIZIO_SP[final]: return None
    dias = ["segunda","terça","quarta","quinta","sexta"]
    for ini, fim in RODIZIO_BLOQUEIO:
        if ini <= hora < fim:
            return f"🚫 Rodízio: placa **{final}** restrita na **{dias[RODIZIO_SP[final]]}** entre {ini.strftime('%H:%M')} e {fim.strftime('%H:%M')}."
    return None

def verificar_conflito_loja(df, data, hora, loja, excluir_idx=-1):
    ds, hs = data.strftime("%d/%m/%Y"), hora.strftime("%H:%M")
    mask = ((df["status"]=="Agendado") & (df["data_entrega"].astype(str)==ds) &
            (df["hora_entrega"].astype(str)==hs) & (df["loja_entrega"].astype(str)==loja))
    if excluir_idx >= 0: mask = mask & (df.index != excluir_idx)
    conf = df[mask]
    if not conf.empty:
        v = conf.iloc[0]
        return f"🚫 Conflito em **{loja}** às **{hs}** — {v.get('modelo','?')} | {v.get('cliente','?')}"
    return None

# ── Cache & envio ─────────────────────────────────────────────────────────────

@st.cache_data(ttl=60)
def gv_carregar():
    try:
        df = pd.read_csv(GV_SHEET_URL, header=0)
        df.columns = [c.strip().lower().replace(" ","_") for c in df.columns]
        for col in ["data_chegada","data_entrega","criado_em","atualizado_em"]:
            if col in df.columns:
                df[col] = df[col].astype(str).apply(fmt_data)
        return df
    except:
        return pd.DataFrame(columns=GV_COLUNAS)

def gv_enviar(payload):
    try:
        requests.post(GV_WEBHOOK, data=json.dumps(payload),
                      headers={"Content-Type":"text/plain"}, timeout=30)
        return True
    except Exception as e:
        st.error(f"Erro ao salvar: {e}"); return False

def gv_novo_id():
    return "VEI" + datetime.datetime.now().strftime("%Y%m%d%H%M%S")

def gv_val_row(row, col, default=""):
    v = row.get(col, default)
    return "" if pd.isna(v) else str(v)

# ── CSS ───────────────────────────────────────────────────────────────────────

GV_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700;800&display=swap');

/* Reset */
section[data-testid="stMain"] * { font-family: 'Montserrat', sans-serif !important; }

/* ── KPIs ── */
.kpi-row { display:flex; gap:10px; margin:14px 0 22px; }
.kpi-box {
    flex:1; background:#fff;
    border:1.5px solid #e8e0d0;
    border-top:4px solid #b57b3f;
    border-radius:14px;
    padding:16px 10px 12px;
    text-align:center;
    box-shadow:0 2px 10px rgba(181,123,63,.07);
    transition:transform .15s, box-shadow .15s;
}
.kpi-box:hover { transform:translateY(-2px); box-shadow:0 6px 18px rgba(181,123,63,.13); }
.kpi-n { font-size:28px; font-weight:800; line-height:1; }
.kpi-l { font-size:10px; color:#94a3b8; margin-top:5px;
          text-transform:uppercase; letter-spacing:.7px; font-weight:600; }

/* ── Card veículo ── */
.vcard {
    background:#fff;
    border:1.5px solid #e8e0d0;
    border-left:5px solid #b57b3f;
    border-radius:13px;
    padding:13px 18px;
    margin-bottom:6px;
    box-shadow:0 1px 5px rgba(0,0,0,.04);
    transition:box-shadow .18s, transform .14s, border-color .18s, background .18s;
}
.vcard:hover {
    border-color:#b57b3f;
    box-shadow:0 5px 18px rgba(181,123,63,.13);
    transform:translateY(-1px);
    background:#fdfbf7;
}
.vcard.ativo {
    background:#fdf8f0;
    border-color:#b57b3f;
    box-shadow:0 5px 22px rgba(181,123,63,.18);
}
.vcard-row   { display:flex; align-items:center; gap:14px; }
.vcard-left  { flex:1; min-width:0; }
.vcard-modelo { font-size:15px; font-weight:700; color:#213144; }
.vcard-fab    { font-size:12px; font-weight:400; color:#94a3b8; margin-left:6px; }
.vcard-tags   { display:flex; flex-wrap:wrap; gap:5px; margin-top:5px; }
.vtag {
    background:#f4f0e8; border:1px solid #e5ddd0;
    border-radius:6px; padding:2px 8px;
    font-size:11px; color:#6b5c45;
}
.vcard-cli  { font-size:11px; color:#475569; margin-top:4px; }
.vcard-right { text-align:right; flex-shrink:0; }
.vbadge { display:inline-block; padding:4px 13px; border-radius:999px;
           font-size:11px; font-weight:700; color:#fff; white-space:nowrap; }
.vidade { font-size:11px; color:#94a3b8; margin-top:5px; }
.vag    { font-size:11px; color:#0891b2; margin-top:2px; }

/* ── Painel de ações (inline) ── */
.painel-wrap {
    background:#fff;
    border:2px solid #b57b3f;
    border-radius:14px;
    padding:22px 26px 18px;
    margin:-2px 0 10px;
    box-shadow:0 8px 28px rgba(181,123,63,.14);
    animation:slideDown .2s ease;
}
@keyframes slideDown {
    from { opacity:0; transform:translateY(-8px); }
    to   { opacity:1; transform:translateY(0); }
}
.p-title { font-size:18px; font-weight:800; color:#213144; }
.p-fab   { font-size:13px; font-weight:400; color:#94a3b8; margin-left:7px; }
.p-sub   { font-size:12px; color:#64748b; margin-top:5px; line-height:1.6; }
.p-badge { display:inline-block; padding:4px 14px; border-radius:999px;
           font-size:11px; font-weight:700; color:#fff; margin-top:8px; }

/* ── Detalhe grid ── */
.det-sec  { margin-bottom:18px; }
.det-head { font-size:10px; font-weight:700; color:#b57b3f;
            text-transform:uppercase; letter-spacing:.8px;
            border-bottom:1px solid #f0e8d8;
            padding-bottom:4px; margin-bottom:10px; }
.det-grid { display:grid;
            grid-template-columns:repeat(auto-fill,minmax(160px,1fr));
            gap:10px 20px; }
.det-l { font-size:10px; color:#94a3b8; font-weight:600; text-transform:uppercase; }
.det-v { font-size:13px; color:#213144; font-weight:500; margin-top:2px; }
.det-v.em { color:#cbd5e1; font-style:italic; }
</style>
"""

# ── Gerador de PDF ───────────────────────────────────────────────────────────

CARD_SHEET_URL = "https://docs.google.com/spreadsheets/d/1RbDl9eD5MafLLQ0QisBy3KQzJOJSGqELyuyGEvZUfm8/export?format=csv&gid=0"

# ── Registro da fonte Montserrat (fonte oficial Carrera Signature) ────────────
_MONTSERRAT_REGISTERED = False

def _registrar_montserrat():
    """Baixa e registra Montserrat no ReportLab. Executado uma vez por sessão."""
    global _MONTSERRAT_REGISTERED
    if _MONTSERRAT_REGISTERED:
        return True
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        import urllib.request, zipfile, io as _io, tempfile, os as _os

        url = "https://fonts.google.com/download?family=Montserrat"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as r:
            data = r.read()

        needed = {
            "Montserrat-Regular":    "Montserrat/static/Montserrat-Regular.ttf",
            "Montserrat-Bold":       "Montserrat/static/Montserrat-Bold.ttf",
            "Montserrat-SemiBold":   "Montserrat/static/Montserrat-SemiBold.ttf",
            "Montserrat-ExtraBold":  "Montserrat/static/Montserrat-ExtraBold.ttf",
        }
        tmpdir = tempfile.mkdtemp()
        with zipfile.ZipFile(_io.BytesIO(data)) as z:
            for font_name, zip_path in needed.items():
                fname = _os.path.join(tmpdir, f"{font_name}.ttf")
                with open(fname, "wb") as f:
                    f.write(z.read(zip_path))
                pdfmetrics.registerFont(TTFont(font_name, fname))

        _MONTSERRAT_REGISTERED = True
        return True
    except Exception:
        return False  # fallback para Helvetica

@st.cache_data(ttl=600, show_spinner="Carregando dados dos segmentos...")
def _load_card_data() -> list:
    """Carrega planilha de cards via pandas. Cache de 10 min."""
    df = pd.read_csv(CARD_SHEET_URL, header=0)
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    return df.fillna("").to_dict(orient="records")

def card_segmentos_disponiveis() -> list:
    try:
        data = _load_card_data()
        return sorted({d.get("segmento","") for d in data if d.get("segmento","")})
    except Exception:
        return ["Sign & Drive", "S&D Empresas", "Nissan Move",
                "AssineCar GWM", "GM Fleet", "GAC Go and Drive", "AssineCar Multbrand"]

def card_lookup(segmento: str, loja: str) -> dict:
    """Busca segmento + loja na planilha. Match case-insensitive."""
    try:
        data = _load_card_data()
    except Exception:
        return {}

    seg_n = segmento.strip().lower()
    loj_n = loja.strip().lower()

    # 1. Match exato (case-insensitive)
    for d in data:
        if d.get("segmento","").strip().lower() == seg_n and \
           d.get("loja","").strip().lower() == loj_n:
            return d

    # 2. Match parcial da loja (ex: "LOJA ALPHAVILLE" → "Loja Alphaville")
    loj_base = loj_n.replace("loja ","").strip()
    for d in data:
        if d.get("segmento","").strip().lower() != seg_n:
            continue
        loja_d_base = d.get("loja","").strip().lower().replace("loja ","").strip()
        if loj_base and loja_d_base and (loj_base in loja_d_base or loja_d_base in loj_base):
            return d

    # 3. Fallback: primeiro do segmento
    for d in data:
        if d.get("segmento","").strip().lower() == seg_n:
            return d

    return data[0] if data else {}





def gerar_pdf_agendamento(row, sv_fn, segmento: str = "") -> bytes:
    """Gera card de agendamento — padrão visual Carrera Signature."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm, mm
        from reportlab.platypus import (Paragraph, Spacer, Table, TableStyle,
                                         HRFlowable, BaseDocTemplate,
                                         Frame, PageTemplate)
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER
        import os
    except ImportError:
        return (f"CARRERA SIGNATURE\nCliente: {sv_fn(row,'cliente')}\n"
                f"Placa: {sv_fn(row,'placa')}\nData: {sv_fn(row,'data_entrega')} "
                f"{sv_fn(row,'hora_entrega')}\nLoja: {sv_fn(row,'loja_entrega')}\n"
                "Instale reportlab: pip install reportlab").encode("utf-8")

    def sv(c): return sv_fn(row, c)

    seg  = segmento if segmento else sv("locadora")
    info = card_lookup(seg, sv("loja_entrega"))
    def inf(k, fb=""): return info.get(k, fb) or fb

    # Tenta registrar Montserrat — usa Helvetica se falhar
    _mont = _registrar_montserrat()
    R  = "Montserrat"         if _mont else "Helvetica"        # Regular
    B  = "Montserrat-Bold"    if _mont else "Helvetica-Bold"   # Bold
    SB = "Montserrat-SemiBold"if _mont else "Helvetica-Bold"   # SemiBold
    XB = "Montserrat-ExtraBold"if _mont else "Helvetica-Bold"  # ExtraBold

    # ── Paleta ──────────────────────────────────────────
    AZUL     = colors.HexColor("#213144")
    AZUL2    = colors.HexColor("#2e4a6b")
    DOURADO  = colors.HexColor("#b57b3f")
    DOURADO2 = colors.HexColor("#dfc28a")
    BRANCO   = colors.white
    CINZA    = colors.HexColor("#475569")
    CINZA_BG = colors.HexColor("#f4f2ef")
    CINZA_BD = colors.HexColor("#ddd8d0")

    W, H   = A4
    MG     = 1.5*cm   # ← MARGEM LATERAL do PDF (esquerda e direita)

    # ── Alturas das seções (calculadas de baixo para cima) ──────────────────
    FOOTER_H  = 1.3*cm  # ← ALTURA da faixa azul do rodapé (telefones)
    REDES_H   = 0.9*cm  # ← ALTURA da faixa de redes sociais (acima do footer)
    AVISO_H   = 0.7*cm  # ← ALTURA da linha de aviso de reagendamento
    BOTTOM_H  = FOOTER_H + REDES_H + AVISO_H + 0.5*cm  # total da área inferior

    HEADER_H = H * 0.24  # ← ALTURA do bloco azul do topo (% da página — aumente para mais azul)
    CARD_H   = 6.0*cm    # ← ALTURA do card branco com os dados do agendamento
    CARD_Y   = H - HEADER_H - 0.5*cm  # posição vertical do topo do card

    BODY_TOP = CARD_Y - CARD_H - 0.6*cm  # onde começa o grid de informações
    BODY_BOT = BOTTOM_H
    CW_BODY  = W - 2*MG  # largura útil do conteúdo (descontando as margens)

    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "..", "LOGO_SIGNATURE.png")

    # ── Canvas: tudo que não é flowable ─────────────────
    def draw_page(c, doc):
        c.saveState()

        # Fundo azul (header)
        c.setFillColor(AZUL)
        c.rect(0, H-HEADER_H, W, HEADER_H, fill=1, stroke=0)

        # Fundo branco (corpo)
        c.setFillColor(BRANCO)
        c.rect(0, BOTTOM_H, W, H-HEADER_H-BOTTOM_H, fill=1, stroke=0)

        # Faixa dourada separadora header/corpo
        c.setFillColor(DOURADO)
        c.rect(0, H-HEADER_H-3, W, 5, fill=1, stroke=0)

        # Pontos decorativos
        c.setFillColor(AZUL2)
        for r in range(6):
            for col in range(6):
                c.circle(W-1.0*cm-col*0.27*cm,
                         H-0.7*cm-r*0.27*cm, 0.038*cm, fill=1, stroke=0)

        # Logo Carrera Signature (PNG local com fundo preto transparente)
        logo_x = MG; logo_y = H-3.0*cm
        logo_w = 4.5*cm; logo_h = 2.2*cm
        if os.path.exists(logo_path):
            c.drawImage(logo_path, logo_x, logo_y,
                        width=logo_w,   # ← LARGURA da logo
                        height=logo_h,  # ← ALTURA da logo
                        preserveAspectRatio=True, mask=[0,30,0,30,0,30])
        else:
            c.setFont(B,16); c.setFillColor(BRANCO)
            c.drawString(logo_x, logo_y+0.5*cm, "Carrera Signature")
        # Logo clicável — abre site da Carrera Signature
        c.linkURL("https://www.carrerasignature.com.br",
                  (logo_x, logo_y, logo_x+logo_w, logo_y+logo_h),
                  relative=0)

        # Título principal do header
        c.setFont(XB, 20)  # ← TAMANHO do título "Chegou a hora..."
        c.setFillColor(BRANCO)
        c.drawString(MG, H-3.8*cm, "Chegou a hora de pegar seu carro!")

        # Subtítulo abaixo do título
        c.setFont(R, 8.5)  # ← TAMANHO do subtítulo dourado
        c.setFillColor(DOURADO2)
        c.drawString(MG, H-4.3*cm,
                     "Aqui estão as informações para a retirada do seu veículo.")

        # ── Card de dados do agendamento ─────────────────
        CX = MG; CY = CARD_Y

        # Sombra do card (deslocamento de 3pt para efeito de profundidade)
        c.setFillColor(colors.HexColor("#bdb5a5"))
        c.roundRect(CX+3, CY-CARD_H-3, CW_BODY, CARD_H, 8, fill=1, stroke=0)
        # Fundo branco do card
        c.setFillColor(BRANCO)
        c.roundRect(CX, CY-CARD_H, CW_BODY, CARD_H, 8, fill=1, stroke=0)
        # Stripe dourada na borda esquerda do card
        c.setFillColor(DOURADO)
        c.roundRect(CX, CY-CARD_H, 6, CARD_H, 3, fill=1, stroke=0)

        # ── Posicionamento interno do card ───────────────
        LX  = CX + 0.9*cm        # ← POSIÇÃO X da coluna esquerda (Cliente, Veículo, Chassi...)
        C2  = CX + CW_BODY*0.52  # ← POSIÇÃO X da coluna direita (Placa, Data, Entregador...)
        TY  = CY - 0.65*cm       # ← POSIÇÃO Y do topo do conteúdo dentro do card
        RH  = 1.08*cm            # ← ESPAÇAMENTO vertical entre cada linha do card
        SZ  = 10                  # ← TAMANHO DA FONTE de todos os valores do card
        FB  = B                   # ← FONTE de todos os valores (Montserrat Bold)

        def lbl(t, x, y):
            c.setFont(B, 7)  # ← TAMANHO dos labels (CLIENTE, PLACA, etc.)
            c.setFillColor(DOURADO)          # cor dourada dos labels
            c.drawString(x, y, t.upper())

        def val(t, x, y):
            c.setFont(FB, SZ)   # FB e SZ definidos acima — mude lá para afetar todos os valores
            c.setFillColor(AZUL)
            c.drawString(x, y, str(t))

        def hsep(y):
            # Linha separadora horizontal entre as linhas do card
            c.setStrokeColor(CINZA_BD); c.setLineWidth(0.4)
            c.line(LX-0.1*cm, y, CX+CW_BODY-0.5*cm, y)

        # Largura máxima de texto por coluna (em pontos) — evita sobreposição
        MAX_L = CW_BODY * 0.48   # ← LARGURA MÁXIMA da coluna esquerda (48% do corpo)
        MAX_R = CW_BODY * 0.46   # ← LARGURA MÁXIMA da coluna direita  (46% do corpo)

        def truncate(txt, max_pts, font=FB, size=SZ):
            """Trunca texto para caber em max_pts pontos de largura."""
            from reportlab.pdfbase.pdfmetrics import stringWidth
            while len(txt) > 1 and stringWidth(txt, font, size) > max_pts:
                txt = txt[:-1]
            return txt if stringWidth(txt, font, size) <= max_pts else txt[:-1] + "…"

        # Linha 1 — Cliente | Placa
        cliente_val = sv("cliente")
        if cliente_val in ("—","Estoque","","nan"): cliente_val = "—"
        lbl("Cliente", LX, TY)
        val(truncate(cliente_val, MAX_L), LX, TY-0.37*cm)
        lbl("Placa",   C2, TY)
        val(sv("placa"), C2, TY-0.37*cm)
        hsep(TY - RH*0.82)

        # Linha 2 — Veículo | Data · Hora
        lbl("Veículo",       LX, TY-RH)
        modelo_txt = f"{sv('modelo')}  ·  {sv('fabricante')}"
        val(truncate(modelo_txt, MAX_L), LX, TY-RH-0.37*cm)
        lbl("Data  ·  Hora", C2, TY-RH)
        val(f"{sv('data_entrega')}   {sv('hora_entrega')}", C2, TY-RH-0.37*cm)
        hsep(TY - RH*1.82)

        # Linha 3 — Chassi | Entregador
        lbl("Chassi",    LX, TY-2*RH)
        val(sv("chassi"), LX, TY-2*RH-0.37*cm)
        lbl("Entregador", C2, TY-2*RH)
        val(sv("entregador") if sv("entregador")!="—" else "—", C2, TY-2*RH-0.37*cm)
        hsep(TY - RH*2.82)

        # Linha 4 — Cor | Consultor
        lbl("Cor",       LX, TY-3*RH)
        val(sv("cor"),   LX, TY-3*RH-0.37*cm)
        lbl("Consultor", C2, TY-3*RH)
        val(sv("consultor") if sv("consultor")!="—" else "—", C2, TY-3*RH-0.37*cm)
        hsep(TY - RH*3.82)

        # Linha 5 — Local de entrega (full width)
        lbl("Local de Entrega", LX, TY-4*RH)
        endereco    = inf("endereco_entrega") or sv("loja_entrega")
        local_txt   = f"{sv('loja_entrega')}   ·   {endereco}"
        val(truncate(local_txt, CW_BODY-1.5*cm), LX, TY-4*RH-0.37*cm)

        # ── Footer azul (full width — de borda a borda) ──────────────────
        FY = 0  # posição Y do footer (começa do zero = base da página)
        c.setFillColor(AZUL)
        c.rect(0, FY, W, FOOTER_H, fill=1, stroke=0)  # retângulo azul full-width

        # Monta pares label + valor dinamicamente
        # Cada par: (label, valor) — o valor vem da planilha do segmento
        pares = [
            ("Central 24h:", c24_v),
            ("Carrera Signature:", csig_v),
            ("Sem Parar:", csem_v),
        ]
        n_pares  = len(pares)
        col_w    = W / n_pares   # divide igualmente entre os pares
        linha_y  = FY + FOOTER_H * 0.6  # linha de cima (label)
        valor_y  = FY + FOOTER_H * 0.18 # linha de baixo (valor)

        for i, (label, valor) in enumerate(pares):
            cx = col_w * i + col_w / 2  # centro de cada coluna
            # Label (branco, menor)
            c.setFont(R, 7); c.setFillColor(BRANCO)
            c.drawCentredString(cx, linha_y, label)
            # Valor (dourado, maior, negrito) — adapta tamanho se texto longo
            from reportlab.pdfbase.pdfmetrics import stringWidth
            sz_val = 9.5
            while sz_val > 6 and stringWidth(valor, B, sz_val) > col_w - 8:
                sz_val -= 0.5
            c.setFont(B, sz_val); c.setFillColor(DOURADO2)
            c.drawCentredString(cx, valor_y, valor)

        # ── Faixa de redes sociais (acima do footer azul) ─────────────────
        RY = FOOTER_H
        c.setFillColor(colors.HexColor("#1a2a3a"))
        c.rect(0, RY, W, REDES_H, fill=1, stroke=0)

        # Ícones embutidos como base64 — sem dependência de rede
        import base64 as _b64, tempfile as _tmp
        _WA = "iVBORw0KGgoAAAANSUhEUgAAAEAAAABACAYAAACqaXHeAAAB9ElEQVR4nO1bQbbCIAxM+/5V6kqvVc+k16orPUxd4asIJQkpJB9m3cLMkNCxlgEKYVrmlXL963IfjuKyxWGTUAWncJQh4oNKC/chbYTYYEcL9yFlRPYgpYX7yDWCfXNt4T64Roycm7SJB+BzIhugUbwDhxu6bDQLDwHbEqgKsCYeAM85aYBF8Q4Y7n9Skz3PN6mhUDg9riLj7PYJxsHSwn1gjNjbD6ItYEE8lsOelqABVsQ75JjACkKaxDtwOf0YkFp9jeIdUtxC2lgV8J/wZYDl1XegVkHzFfAJQrmJTyqYYJFTjdMyry4biFRAafGSc44AtvM+F05z83tAN6A2gdoYJfq/Rj6QmHNa5tXs+wApNN8CYhVgKQht0YOQyCiG0Q2oTaA2ugESg1gNQgCG/xiRQm+BUl9jacTrch96BVAurpH4qKBy7BUAQPvASHMVULhlvRXWaAKX08cA6tNAkwlULlutWXuABhNyOXwZYCkTnB5Xlnhfo1gUDkFDhaTw0wKUKtjL/xrFh7SJV4BG4XsIboKYKgitvmbxMU3RpwA1HFkUD5B4DMZu3K6+ZuEA6YVk7wHahWORDEK+g8/zzYx4TBs3/7k8OflpN4KaZsm/BTTHZQ431o8hjSZwOfVjc1JEmj046aPZo7M+mj08HYPW4/Nv3G7lCi6ODTQAAAAASUVORK5CYII="
        _IG = "iVBORw0KGgoAAAANSUhEUgAAAEAAAABACAYAAACqaXHeAAACCElEQVR4nO2bPZLCMAyFlcxeCUo4GJRwMCjhUGylnaD1jyTbiuL4VQzEQd+zLBkmnsBIj8PtI7n+/L5OrWJZqtmXSIFzamVI9ZvWBqeqbUS1m7UGp6plRPFNrMGpSo1QD14bnEprxKwZ5A0eQB+T2ACP8ChNbOy08QweEndJsDJga/AA/JizBmwRHsWJXVUEe1LSgC3PPirHEDWgB3hUiiVoQE/wqBjTqAH0jR5nHxVi+ym54el1KRleTc/jXT32KwMks+8FHkAWC2VU1QBP8ChtTH8GcGffIzyKG9uSdXQBgL4rf0zIvPsMKGqDWsXa1hr1xdSAXL/Gzy2NmK3Wv2SzUrKxkehxuH1MaoAGyMqE5ksgBBJLcXrt83hvvhzMu0AKaI0i2NQAOqMcQHpN66VglgGS2TXtAmbf5FRmBnhsgwCNDdCsZ03dKJH5EkiZYDnzqOb7gNPrEuzv3LGtZZIBGhCrTjBbPY3lsQ2e39fJ9Ncggu325zDK0/+KYyMEYPdUpich88gAfMHNgjU2K1xxY1uyqjLAownamL4MkNQCTyZIYqGMRW3Qkwla/VsCPXeEENvoAqE3e8yCGFM0A3oyIcWSXAI9mJBjGDUgd8GWs4ATOysDtmgCN2YxmPenSaSTJa4BnrNBE5uqCHo0QRvTODZXK5DdHpyk2u3RWardHp6Oyevx+V/9b+oBw/MqvwAAAABJRU5ErkJggg=="
        _YT = "iVBORw0KGgoAAAANSUhEUgAAAEAAAABACAYAAACqaXHeAAAB2klEQVR4nO2b25KDIBBEG2v//5fdlx2LsCIX59Jo+jEVcfrQg6QiCU7agX3k+wlIVrUU97HRqOGWrICoD6ptvJQ2CLXBrI2X0gJxexBv46Xugpi+ONp4qVkQ28xFbOaB+ZqGATCaF83U1h0bZuNn6m2JrgSsZh7or7kJYEXzop7apxbBJ+kSwMqzL2p5qAJ4gnnRlZdTAE8yL6p5+q4B5Qfqs7/zhOnMm08CiCCU+gBg2vv7TgGi9Oi/BhBAyHUAcF35gyHkXuOeAiQtsQHBz/0gCOKZYx8QmAYOAKIACFwAAHcIybz/7xhK9n8O8SUgl0MauAEA5hD4AQCmT4k1AIgMIKwFAFBPw3oAREoQ1gWg9Ij8URnFU8p7g7USYLAxWiMBhjtC/gQYb4c3r7expmRsPgGJswUcfgSJ+FrA0TzABsDZPPAHIHwdSMndvHiOT0DArOc6ALinIGDWj1tnXmMSEDzruT4AuKQg2Hzp0W8fQDTruf61gEkKSMyfebNfA0jM13QKIHxfYKCap2oCngThystlCzwBQstD/E4wWE0AK6egp/auBKwIobfmYWPsb5GOTtbwGsCchpnaphZBRgizNX2PzWkV8tqDk6Vee3S21GsPT9fEenz+F4WldHkN6+BcAAAAAElFTkSuQmCC"

        redes = [
            (_WA,  "WhatsApp",  "Carrera Signature", "https://wa.me/551140037214"),
            (_IG,  "Instagram", "@carrerasignature",  "https://www.instagram.com/carrerasignature"),
            (_YT,  "YouTube",   "Carrera Signature",  "https://www.youtube.com/@carrerasignature"),
        ]
        n_redes = len(redes)
        rc_w    = W / n_redes
        icon_sz = 0.38*cm   # ← TAMANHO dos ícones das redes
        ry_mid  = RY + REDES_H / 2

        for i, (b64_str, rede, handle, url) in enumerate(redes):
            cx  = rc_w * i + rc_w / 2
            rx0 = rc_w * i; rx1 = rx0 + rc_w

            # Escreve PNG temporário a partir do base64
            try:
                icon_data = _b64.b64decode(b64_str)
                with _tmp.NamedTemporaryFile(suffix=".png", delete=False) as tf:
                    tf.write(icon_data); tf_path = tf.name
                # Calcula posição: ícone + texto centralizados juntos na coluna
                gap     = 0.1*cm
                from reportlab.pdfbase.pdfmetrics import stringWidth as _sw
                txt_w   = max(_sw(rede, B, 7.5),
                              _sw(handle, R, 6.5))
                total_w = icon_sz + gap + txt_w
                ix      = cx - total_w / 2
                tx      = ix + icon_sz + gap
                icon_y  = ry_mid - icon_sz / 2 + 0.02*cm
                c.drawImage(tf_path, ix, icon_y,
                            width=icon_sz, height=icon_sz,
                            preserveAspectRatio=True, mask="auto")
                os.unlink(tf_path)
                c.setFont(B, 7.5); c.setFillColor(BRANCO)
                c.drawString(tx, ry_mid + 0.07*cm, rede)
                c.setFont(R, 6.5); c.setFillColor(DOURADO2)
                c.drawString(tx, ry_mid - 0.17*cm, handle)
            except Exception:
                # Fallback: só texto centralizado
                c.setFont(B, 7.5); c.setFillColor(BRANCO)
                c.drawCentredString(cx, ry_mid + 0.08*cm, rede)
                c.setFont(R, 6.5); c.setFillColor(DOURADO2)
                c.drawCentredString(cx, ry_mid - 0.18*cm, handle)

            # Área clicável
            c.linkURL(url, (rx0+2, RY+2, rx1-2, RY+REDES_H-2), relative=0)

        # ── Aviso de reagendamento (acima das redes) ──────────────────────
        aviso_y = RY + REDES_H + 0.18*cm
        c.setFont(B, 7.5)  # ← TAMANHO do aviso de reagendamento
        c.setFillColor(AZUL)
        c.drawCentredString(W/2, aviso_y, aviso_v)

        # Linha dourada separando aviso do grid informativo
        c.setStrokeColor(DOURADO2); c.setLineWidth(1.0)
        c.line(MG, aviso_y + 0.42*cm, W-MG, aviso_y + 0.42*cm)

        c.restoreState()

    # Variáveis que o canvas precisa acessar
    c24_v  = inf("central_24hrs",    "0800 071 8090")
    csig_v = inf("carrera_signature","4003-7214")
    csem_v = inf("central_sem_parar","0800 724 2467")
    aviso_v= inf("aviso_reagendamento",
                 "Atrasos maiores que 30 min para a retirada estão sujeitos a reagendamento")

    # ── Estilos do grid de informações (Importante, SAC, etc.) ─────────────
    def sty(name, **kw):
        base = dict(fontName=R, fontSize=9, textColor=AZUL,
                    leading=14, spaceAfter=0)
        base.update(kw); return ParagraphStyle(name, **base)

    st_sl = sty("sl",
                fontName=B,
                fontSize=8.5,        # ← TAMANHO do título de cada card (✓ Importante, ✓ SAC...)
                textColor=DOURADO,   # cor dourada dos títulos
                spaceAfter=4,        # espaço abaixo do título antes do texto
                leading=11)

    st_sb = sty("sb",
                fontSize=8.5,        # ← TAMANHO do texto de cada card informativo
                textColor=CINZA,     # cor cinza do corpo do texto
                leading=13)          # ← ESPAÇAMENTO entre linhas do texto (leading)

    st_rd = sty("rd",
                fontSize=7.5,        # ← TAMANHO da linha "Carrera Signature · Gerado em"
                textColor=CINZA,
                alignment=TA_CENTER)

    HW = CW_BODY / 2  # ← LARGURA de cada coluna do grid (metade do corpo)

    # ── Grid informativo — altura automática, sem corte ──
    def info_cell(titulo, corpo):
        return [
            Paragraph(f"✓  {titulo}", st_sl),
            Paragraph(corpo or "—",   st_sb),
        ]

    # Monta como tabela única de 2 colunas com linhas separadas por célula
    grid_rows = [
        [info_cell("Importante",   inf("importante")),
         info_cell("Combustível",  inf("combustivel"))],
        [info_cell("Sem Parar",    inf("sem_parar")),
         info_cell("SAC",          inf("sac"))],
        [info_cell("Revisões",     inf("revisoes")),
         info_cell("Local Revisão",inf("local_revisao"))],
    ]

    # Cada info_cell retorna uma lista de Paragraphs — precisa ser uma Table interna
    def make_cell(titulo, corpo):
        return Table(
            [[Paragraph(f"✓  {titulo}", st_sl)],
             [Paragraph(corpo or "—",   st_sb)]],
            colWidths=[HW - 0.6*cm]
        )

    grid = Table([
        [make_cell("Importante",   inf("importante")),
         make_cell("Combustível",  inf("combustivel"))],
        [make_cell("Sem Parar",    inf("sem_parar")),
         make_cell("SAC",          inf("sac"))],
        [make_cell("Revisões",     inf("revisoes")),
         make_cell("Local Revisão",inf("local_revisao"))],
    ], colWidths=[HW, HW])
    grid.setStyle(TableStyle([
        ("VALIGN",       (0,0),(-1,-1), "TOP"),
        ("BACKGROUND",   (0,0),(-1,-1), CINZA_BG),
        ("BOX",          (0,0),(-1,-1), 0.5, CINZA_BD),
        ("INNERGRID",    (0,0),(-1,-1), 0.5, CINZA_BD),
        ("LEFTPADDING",  (0,0),(-1,-1), 12),
        ("RIGHTPADDING", (0,0),(-1,-1), 12),
        ("TOPPADDING",   (0,0),(-1,-1), 10),
        ("BOTTOMPADDING",(0,0),(-1,-1), 10),
    ]))

    story = [
        grid,
        Spacer(1, 0.4*cm),
        Paragraph(
            f"Carrera Signature  ·  Gerado em "
            f"{datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}",
            st_rd
        ),
    ]

    buf = io.BytesIO()

    class CardDoc(BaseDocTemplate):
        def __init__(self, bf, **kw):
            super().__init__(bf, **kw)
            fy = BODY_BOT
            fh = BODY_TOP - BODY_BOT
            fr = Frame(MG, fy, CW_BODY, fh, id="b",
                       leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)
            self.addPageTemplates([PageTemplate(id="p", frames=[fr], onPage=draw_page)])

    CardDoc(buf, pagesize=A4).build(story)
    return buf.getvalue()


# ── RENDER ────────────────────────────────────────────────────────────────────

def render():
    st.markdown(GV_CSS, unsafe_allow_html=True)

    autenticado = st.session_state.get("auth_tipo","") == "Staff"

    # Staff com acesso somente leitura (sem edição, cadastro ou deleção)
    STAFF_READONLY = {
        "Andrea Bettega Pereira da Costa",
        "Raymond Jose Duque Bello",
    }
    auth_nome   = st.session_state.get("auth_nome","")
    pode_editar = autenticado and auth_nome not in STAFF_READONLY
    hoje = datetime.date.today()

    # Session state defaults
    for k, v in [("gv_sel", None), ("ag_forcar", False), ("gv_cad_open", False)]:
        if k not in st.session_state:
            st.session_state[k] = v

    # Carrega dados
    df_gv = gv_carregar()
    if not df_gv.empty:
        df_gv["_idade"] = df_gv.apply(calcular_idade, axis=1)

    # ── Helper local ─────────────────────────────────────
    def sv(r, c):
        v = r.get(c, ""); s = str(v).strip()
        if s in ("","nan","None","NaT"): return "—"
        if s.endswith(".0") and s[:-2].lstrip("-").isdigit(): return s[:-2]
        return s

    # ── Header ───────────────────────────────────────────
    h1, h2, h3, h4 = st.columns([5, 1, 1, 1])
    with h1:
        st.markdown(
            f"<h2 style='color:{AZUL};margin:0;font-family:Inter,sans-serif'>"
            f"🚘 Estoque de Veículos</h2>",
            unsafe_allow_html=True,
        )
    with h2:
        if st.button("🔄 Atualizar", use_container_width=True, key="gv_ref"):
            gv_carregar.clear(); st.rerun()
    with h3:
        if pode_editar and st.button("➕ Cadastrar", use_container_width=True,
                                      key="gv_cad_btn", type="primary"):
            st.session_state["gv_cad_open"] = not st.session_state["gv_cad_open"]
            st.session_state["gv_sel"] = None
            st.rerun()
    with h4:
        if autenticado and not df_gv.empty:
            try:
                buf = io.BytesIO()
                df_gv.to_excel(buf, index=False, engine="openpyxl"); buf.seek(0)
                st.download_button(
                    "📥 Excel", data=buf.getvalue(),
                    file_name="veiculos.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, key="gv_xl",
                )
            except:
                st.download_button(
                    "📥 CSV", data=df_gv.to_csv(index=False).encode(),
                    file_name="veiculos.csv", mime="text/csv",
                    use_container_width=True, key="gv_csv",
                )

    if df_gv.empty:
        st.info("Nenhum veículo cadastrado ainda.")
        return

    # ── Navegação principal ──────────────────────────────
    if pode_editar:
        aba_principal = st.tabs(["🚘 Estoque", "📅 Agenda", "📦 Recebimento", "✉️ E-mail das 11"])
    else:
        aba_principal = st.tabs(["🚘 Estoque", "📅 Agenda"])

    with aba_principal[0]:

        # ── Filtros de data ───────────────────────────────────
        primeiro_dia_mes = hoje.replace(day=1)
        fd1, fd2, fd3 = st.columns([2, 2, 5])
        with fd1:
            data_ini_est = st.date_input("📅 Data início", value=primeiro_dia_mes, key="est_di", format="DD/MM/YYYY")
        with fd2:
            data_fim_est = st.date_input("📅 Data fim",    value=hoje,             key="est_df", format="DD/MM/YYYY")

        # Filtra entregas no período para KPIs de entregue/atrasado
        datas_periodo_est = {
            (data_ini_est + datetime.timedelta(days=d)).strftime("%d/%m/%Y")
            for d in range((data_fim_est - data_ini_est).days + 1)
        }

        # ── KPIs (sempre sobre df_gv completo, data só afeta Entregues) ──
        total     = len(df_gv)
        livres    = len(df_gv[df_gv["status"] == "Livre"])
        trans_l   = len(df_gv[df_gv["status"] == "Trânsito Livre"])
        trans_v   = len(df_gv[df_gv["status"] == "Trânsito Vendido"])
        ag_atr = len(df_gv[df_gv["status"] == "Aguardando Atribuição"])
        ag_ag     = len(df_gv[df_gv["status"] == "Aguardando Agendamento"])
        agend     = len(df_gv[df_gv["status"] == "Agendado"])
        avar      = len(df_gv[df_gv["status"] == "Avariado"])

        # Entregues no período selecionado
        df_entr_periodo = df_gv[
            (df_gv["status"] == "Entregue") &
            (df_gv["data_entrega"].astype(str).isin(datas_periodo_est))
        ]
        entr = len(df_entr_periodo)

        # Atrasados (agendados com farol vermelho)
        ach = df_gv[df_gv["status"] == "Agendado"].copy()
        if not ach.empty:
            ach["_f"] = ach.apply(farol_agendamento, axis=1)
            atras = len(ach[ach["_f"] == "🔴"])
        else:
            atras = 0

        st.markdown(f"""
        <div class="kpi-row">
          <div class="kpi-box">
            <div class="kpi-n" style="color:{AZUL}">{total}</div>
            <div class="kpi-l">Total</div>
          </div>
          <div class="kpi-box">
            <div class="kpi-n" style="color:#22c55e">{livres}</div>
            <div class="kpi-l">Livres</div>
          </div>
          <div class="kpi-box">
            <div class="kpi-n" style="color:#3b82f6">{trans_l}</div>
            <div class="kpi-l">Trânsito Livre</div>
          </div>
          <div class="kpi-box">
            <div class="kpi-n" style="color:#8b5cf6">{trans_v}</div>
            <div class="kpi-l">Trânsito Vendido</div>
          </div>
          <div class="kpi-box">
            <div class="kpi-n" style="color:#8b5cf6">{ag_atr}</div>
            <div class="kpi-l">Aguardando Atribuição</div>
          </div>
          <div class="kpi-box">
            <div class="kpi-n" style="color:#f97316">{ag_ag}</div>
            <div class="kpi-l">Ag. Agendamento</div>
          </div>
          <div class="kpi-box">
            <div class="kpi-n" style="color:#06b6d4">{agend}</div>
            <div class="kpi-l">Agendados</div>
          </div>
          <div class="kpi-box">
            <div class="kpi-n" style="color:#10b981">{entr}</div>
            <div class="kpi-l">Entregues</div>
          </div>
          <div class="kpi-box">
            <div class="kpi-n" style="color:#ef4444">{avar}</div>
            <div class="kpi-l">Avariados</div>
          </div>
          <div class="kpi-box">
            <div class="kpi-n" style="color:#dc2626">{atras}</div>
            <div class="kpi-l">Atrasados</div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        if data_ini_est != primeiro_dia_mes or data_fim_est != hoje:
            st.caption(
                f"📅 Entregues filtrados: {data_ini_est.strftime('%d/%m/%Y')} a {data_fim_est.strftime('%d/%m/%Y')}"
            )

        # ── Filtros ──────────────────────────────────────────
        f1, f2, f3, f4 = st.columns(4)
        with f1: flt_sta = st.selectbox("Status", ["Todos"] + GV_STATUS_LIST, key="p_sta")
        with f2:
            fabs = ["Todos"] + sorted(df_gv["fabricante"].dropna().unique()) if "fabricante" in df_gv.columns else ["Todos"]
            flt_fab = st.selectbox("Fabricante", fabs, key="p_fab")
        with f3:
            locs = ["Todos"] + sorted(df_gv["locadora"].dropna().unique()) if "locadora" in df_gv.columns else ["Todos"]
            flt_loc = st.selectbox("Locadora", locs, key="p_loc")
        with f4:
            cons = ["Todos"] + sorted(df_gv["consultor"].dropna().unique()) if "consultor" in df_gv.columns else ["Todos"]
            flt_con = st.selectbox("Consultor", cons, key="p_con")

        b1, b2, b3 = st.columns(3)
        with b1: s_ch = st.text_input("🔑 Chassi",    placeholder="Chassi", key="b_ch")
        with b2: s_pl = st.text_input("🪪 Placa",     placeholder="Placa",  key="b_pl")
        with b3: s_pe = st.text_input("📄 Nº Pedido", placeholder="Nº Pedido",    key="b_pe")

        dv = df_gv.copy()
        if flt_sta != "Todos": dv = dv[dv["status"]     == flt_sta]
        if flt_fab != "Todos": dv = dv[dv["fabricante"] == flt_fab]
        if flt_loc != "Todos": dv = dv[dv["locadora"]   == flt_loc]
        if flt_con != "Todos": dv = dv[dv["consultor"]  == flt_con]
        if s_ch: dv = dv[dv["chassi"].astype(str).str.lower().str.contains(s_ch.lower(), na=False)]
        if s_pl: dv = dv[dv["placa"].astype(str).str.lower().str.contains(s_pl.lower(),  na=False)]
        if s_pe: dv = dv[dv["pedido"].astype(str).str.lower().str.contains(s_pe.lower(), na=False)]

        st.markdown(
            f"<p style='color:#94a3b8;font-size:13px;margin:6px 0 10px'>"
            f"<b style='color:{AZUL}'>{len(dv)}</b> veículo(s)</p>",
            unsafe_allow_html=True,
        )
        st.divider()

        # ══════════════════════════════════════════════════════
        # PAINEL CADASTRAR
        # ══════════════════════════════════════════════════════
        if st.session_state.get("gv_cad_open") and pode_editar:
            with st.container(border=True):
                cx1, cx2 = st.columns([6, 1])
                with cx1:
                    st.markdown(
                        f"<div style='font-size:17px;font-weight:800;color:{AZUL}'>"
                        f"➕ Cadastrar Veículo</div>",
                        unsafe_allow_html=True,
                    )
                with cx2:
                    if st.button("✕ Fechar", key="cad_x", use_container_width=True):
                        st.session_state["gv_cad_open"] = False; st.rerun()

                modo = st.radio("", ["✏️ Manual", "📂 Importar Excel"], horizontal=True, key="cad_modo")

                # ── Manual ────────────────────────────────
                if "Manual" in modo:
                    with st.form("f_cad"):
 
                        # ── Identificação ────────────────────────────
                        st.markdown("**🔍 Identificação**")
                        c1, c2, c3 = st.columns(3)
                        with c1:
                            fab    = st.selectbox("Fabricante *", GV_FABRICANTES)
                            mod    = st.text_input("Modelo *")
                            cor    = st.text_input("Cor *")
                        with c2:
                            chassi = st.text_input("Chassi *")
                            placa  = st.text_input("Placa")
                            comb   = st.selectbox("Combustível", GV_COMBUSTIVEIS)
                        with c3:
                            anof   = st.text_input("Ano Fabricação")
                            anom   = st.text_input("Ano Modelo")
                            opc    = st.text_input("Opcionais")
 
                        # ── Operacional ──────────────────────────────
                        st.markdown("**🏢 Operacional**")
                        o1, o2, o3 = st.columns(3)
                        with o1:
                            loc    = st.selectbox("Locadora", GV_LOCADORAS)
                            consc  = st.text_input("Consultor")
                            stc    = st.selectbox("Status Inicial", GV_STATUS_LIST)
                        with o2:
                            clic   = st.text_input("Cliente")
                            pedido = st.text_input("Nº Pedido")
                            localc = st.text_input("Local Atual")
                        with o3:
                            av     = st.selectbox("Avaria?", ["Não", "Sim"])
                            obs    = st.text_area("Obs. Avaria", height=68)
 
                        # ── Datas e Entrega ──────────────────────────
                        st.markdown("**📅 Datas e Entrega**")
                        d1, d2, d3, d4 = st.columns(4)
                        with d1:
                            dch  = st.date_input("Data Chegada", value=None, format="DD/MM/YYYY")
                        with d2:
                            det  = st.date_input("Data Entrega", value=None, format="DD/MM/YYYY")
                        with d3:
                            hent = st.time_input("Hora Entrega", value=datetime.time(10, 0))
                        with d4:
                            entregador = st.text_input("Entregador")
                        lj = st.selectbox("Loja de Entrega", [""] + GV_LOJAS)
 
                        # ── Financeiro ───────────────────────────────
                        st.markdown("**💰 Financeiro**")
                        fn1, fn2, fn3 = st.columns(3)
                        with fn1: vnf = st.number_input("Valor NF (R$)", min_value=0.0, step=100.0, value=0.0)
                        with fn2: mgm = st.number_input("Margem (%)",    min_value=0.0, step=0.1,   value=0.0)
                        with fn3: com = st.number_input("Comissão (%)",  min_value=0.0, step=0.1,   value=0.0)
 
                        if st.form_submit_button("💾 Cadastrar", use_container_width=True, type="primary"):
                            if not mod or not chassi:
                                st.error("Modelo e Chassi são obrigatórios.")
                            elif "chassi" in df_gv.columns and chassi.upper() in df_gv["chassi"].astype(str).str.upper().values:
                                st.error("Chassi já existe!")
                            else:
                                agr = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                                nl  = [
                                    gv_novo_id(), fab, mod, chassi, placa, cor,
                                    anof, anom, comb, opc,
                                    loc, consc, clic, pedido, stc, localc,
                                    dch.strftime("%d/%m/%Y") if dch else "",
                                    det.strftime("%d/%m/%Y") if det else "",
                                    hent.strftime("%H:%M"),
                                    entregador,
                                    av, obs if av == "Sim" else "",
                                    lj, vnf, mgm, com,
                                    agr, agr,
                                    st.session_state.get("auth_nome", "Sistema"),
                                    "",  # transporte_solicitado
                                ]
                                pg = st.progress(0, "Salvando..."); pg.progress(70, "Enviando...")
                                if gv_enviar({"aba": "veiculos", "acao": "inserir", "linha": nl}):
                                    pg.progress(100, "Concluído!")
                                    gv_carregar.clear()
                                    st.success("✅ Veículo cadastrado!")
                                    st.session_state["gv_cad_open"] = False
                                    st.rerun()
 

                # ── Importar Excel ────────────────────────
               # ── Importar Excel ────────────────────────
                else:
                    # Colunas completas (todos os 30 campos da base)
                    CABECALHO_XLS = [
                            "FABRICANTE","MODELO","CHASSI","PLACA","COR",
                            "ANO FABRICACAO","ANO MODELO","COMBUSTIVEL","OPCIONAIS",
                            "LOCADORA","CONSULTOR","CLIENTE","PEDIDO","STATUS",
                            "LOCAL ATUAL","DATA CHEGADA","DATA ENTREGA","HORA ENTREGA",
                            "ENTREGADOR","AVARIA","OBS AVARIA","LOJA ENTREGA",
                            "VALOR NF","MARGEM","COMISSAO",
                    ]
                    try:
                        import openpyxl
                        from openpyxl.styles import Font, PatternFill, Alignment
                        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Importação"
                        ws.append(CABECALHO_XLS)
                        ws.append([

                        ])
                        for cell in ws[1]:
                            cell.font      = Font(bold=True, color="FFFFFF")
                            cell.fill      = PatternFill("solid", fgColor="213144")
                            cell.alignment = Alignment(horizontal="center")
                        for col in ws.columns:
                            ws.column_dimensions[col[0].column_letter].width = 20
                        buf_mod = io.BytesIO(); wb.save(buf_mod); buf_mod.seek(0)
                        st.download_button(
                            "📋 Baixar modelo Excel", data=buf_mod.getvalue(),
                            file_name="modelo_importacao_veiculos.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True, key="gv_mod_dl",
                        )
                    except Exception as ex:
                        st.info(f"Erro ao gerar modelo: {ex}")

                    arq = st.file_uploader("Selecione o .xlsx", type=["xlsx"], key="gv_up")
                    if arq:
                        try:
                            di = pd.read_excel(arq)
                            st.dataframe(di.head(5), use_container_width=True)
                            if st.button("📤 Importar", use_container_width=True, key="gv_imp", type="primary"):
                                agr = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                                cex = set(df_gv["chassi"].astype(str).str.upper()) if "chassi" in df_gv.columns else set()
                                # Mapeamento completo de todos os campos
                                col_map = {
                                    "FABRICANTE":     ["FABRICANTE","fabricante","Fabricante"],
                                    "MODELO":         ["MODELO","modelo","Modelo"],
                                    "CHASSI":         ["CHASSI","chassi","Chassi"],
                                    "PLACA":          ["PLACA","placa","Placa"],
                                    "COR":            ["COR","cor","Cor"],
                                    "ANO FABRICACAO": ["ANO FABRICACAO","ANO_FABRICACAO","ano_fabricacao","ANO FAB.","ANO FAB"],
                                    "ANO MODELO":     ["ANO MODELO","ANO_MODELO","ano_modelo","ANO MOD.","ANO MOD"],
                                    "COMBUSTIVEL":    ["COMBUSTIVEL","combustivel","Combustível","COMBUSTÍVEL"],
                                    "OPCIONAIS":      ["OPCIONAIS","opcionais","Opcionais"],
                                    "LOCADORA":       ["LOCADORA","locadora","Locadora"],
                                    "CONSULTOR":      ["CONSULTOR","consultor","Consultor"],
                                    "CLIENTE":        ["CLIENTE","cliente","Cliente"],
                                    "PEDIDO":         ["PEDIDO","pedido","Pedido"],
                                    "STATUS":         ["STATUS","status","Status"],
                                    "LOCAL ATUAL":    ["LOCAL ATUAL","LOCAL_ATUAL","local_atual","Local Atual"],
                                    "DATA CHEGADA":   ["DATA CHEGADA","DATA_CHEGADA","data_chegada","Data Chegada"],
                                    "DATA ENTREGA":   ["DATA ENTREGA","DATA_ENTREGA","data_entrega","Data Entrega"],
                                    "HORA ENTREGA":   ["HORA ENTREGA","HORA_ENTREGA","hora_entrega","Hora Entrega"],
                                    "ENTREGADOR":     ["ENTREGADOR","entregador","Entregador"],
                                    "AVARIA":         ["AVARIA","avaria","COM AVARIA?","COM AVARIA"],
                                    "OBS AVARIA":     ["OBS AVARIA","OBS_AVARIA","obs_avaria","Obs Avaria","Obs. Avaria"],
                                    "LOJA ENTREGA":   ["LOJA ENTREGA","LOJA_ENTREGA","loja_entrega","Loja Entrega","LOJA DE ENTREGA"],
                                    "VALOR NF":       ["VALOR NF","VALOR_NF","valor_nf","Valor NF","Valor Nota"],
                                    "MARGEM":         ["MARGEM","margem","Margem"],
                                    "COMISSAO":       ["COMISSAO","COMISSÃO","comissao","comissão","Comissão"],
                                }
                                errs = 0; dps = []
                                pg2 = st.progress(0, "Importando...")
                                for ri, (_, row) in enumerate(di.iterrows()):
                                    pg2.progress(int(ri/len(di)*90)+5, f"Linha {ri+1}/{len(di)}...")
                                    def g(key):
                                        for alias in col_map.get(key, [key]):
                                            vv = row.get(alias, None)
                                            if vv is not None and not (isinstance(vv, float) and pd.isna(vv)):
                                                return str(vv).strip()
                                        return ""
                                    ci = g("CHASSI").upper()
                                    if ci and ci in cex: dps.append(ci); continue
                                    if ci: cex.add(ci)
                                    av_raw = g("AVARIA") or "Não"
                                    av_val = "Sim" if av_raw.lower() in ("sim","yes","1","true") else "Não"
                                    # Monta nl com todos os 30 campos na ordem exata de GV_COLUNAS
                                    nl = [
                                        gv_novo_id(),       # id
                                        g("FABRICANTE"),    # fabricante
                                        g("MODELO"),        # modelo
                                        ci,                 # chassi
                                        g("PLACA"),         # placa
                                        g("COR"),           # cor
                                        g("ANO FABRICACAO"),# ano_fabricacao
                                        g("ANO MODELO"),    # ano_modelo
                                        g("COMBUSTIVEL"),   # combustivel
                                        g("OPCIONAIS"),     # opcionais
                                        g("LOCADORA"),      # locadora
                                        g("CONSULTOR"),     # consultor
                                        g("CLIENTE"),       # cliente
                                        g("PEDIDO"),        # pedido
                                        g("STATUS") or "Trânsito Livre",  # status
                                        g("LOCAL ATUAL"),   # local_atual
                                        g("DATA CHEGADA"),  # data_chegada
                                        g("DATA ENTREGA"),  # data_entrega
                                        g("HORA ENTREGA"),  # hora_entrega
                                        g("ENTREGADOR"),    # entregador
                                        av_val,             # avaria
                                        g("OBS AVARIA") if av_val=="Sim" else "",  # obs_avaria
                                        g("LOJA ENTREGA"),  # loja_entrega
                                        g("VALOR NF"),      # valor_nf
                                        g("MARGEM"),        # margem
                                        g("COMISSAO"),      # comissao
                                        agr,                # criado_em
                                        agr,                # atualizado_em
                                        "Importação",       # atualizado_por
                                        "",                 # transporte_solicitado
                                    ]
                                    if not gv_enviar({"aba":"veiculos","acao":"inserir","linha":nl}):
                                        errs += 1
                                pg2.progress(100, "Concluído!")
                                gv_carregar.clear()
                                st.success(f"✅ {len(di)-errs-len(dps)} importado(s)!")
                                if dps: st.warning(f"Duplicados ignorados: {', '.join(dps)}")
                                if errs: st.error(f"Erros ao enviar: {errs}")
                                st.session_state["gv_cad_open"] = False
                                st.rerun()
                        except Exception as e:
                            st.error(f"Erro ao ler arquivo: {e}")

        # ══════════════════════════════════════════════════════
        # LISTA DE CARDS
        # ══════════════════════════════════════════════════════
        for i, (_, row) in enumerate(dv.iterrows()):
            ch_r   = sv(row, "chassi")
            st_r   = str(row.get("status", ""))
            cor_r  = STATUS_CORES.get(st_r, "#94a3b8")
            id_r   = row.get("_idade", None)
            id_txt = f"{farol_idade(id_r)} {id_r}d" if id_r is not None else "⚪ —"
            ativo  = st.session_state.get("gv_sel") == ch_r

            cli_h = (
                f"<div class='vcard-cli'>👤 <b>{sv(row,'cliente')}</b> · {sv(row,'consultor')}</div>"
                if sv(row, "cliente") != "—" else ""
            )
            ag_h = (
                f"<div class='vag'>📅 {sv(row,'data_entrega')} {sv(row,'hora_entrega')} · {sv(row,'loja_entrega')}</div>"
                if st_r == "Agendado" and sv(row, "data_entrega") != "—" else ""
            )

            # Card + botão lado a lado
            col_card, col_btn = st.columns([11, 1])
            with col_card:
                st.markdown(
                    f"<div class='vcard {'ativo' if ativo else ''}'>"
                    f"  <div class='vcard-row'>"
                    f"    <div class='vcard-left'>"
                    f"      <div class='vcard-modelo'>{sv(row,'modelo')}"
                    f"        <span class='vcard-fab'>{sv(row,'fabricante')}</span></div>"
                    f"      <div class='vcard-tags'>"
                    f"        <span class='vtag'>🔑 {sv(row,'chassi')}</span>"
                    f"        <span class='vtag'>🪪 {sv(row,'placa')}</span>"
                    f"        <span class='vtag'>🎨 {sv(row,'cor')}</span>"
                    f"        <span class='vtag'>🏢 {sv(row,'locadora')}</span>"
                    + (f"        <span class='vtag'>📄 {sv(row,'pedido')}</span>" if sv(row,'pedido') != '—' else "") +
                    f"      </div>{cli_h}"
                    f"    </div>"
                    f"    <div class='vcard-right'>"
                    f"      <span class='vbadge' style='background:{cor_r}'>{st_r}</span>"
                    f"      <div class='vidade'>{id_txt}</div>{ag_h}"
                    f"    </div>"
                    f"  </div>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
            with col_btn:
                if st.button(
                    "✕" if ativo else "➕",
                    key=f"btn_{i}_{ch_r}",
                    use_container_width=True,
                    type="primary" if ativo else "secondary",
                ):
                    st.session_state["gv_sel"]      = None if ativo else ch_r
                    st.session_state["gv_cad_open"] = False
                    st.rerun()

            # ── Painel inline ──────────────────────────────
            if ativo:
                vm     = df_gv[df_gv["chassi"].astype(str) == ch_r].iloc[0]
                idx_vm = df_gv[df_gv["chassi"].astype(str) == ch_r].index[0]
                lvm    = int(idx_vm) + 2
                cor_p  = STATUS_CORES.get(sv(vm, "status"), "#94a3b8")
                id_p   = vm.get("_idade", None)

                with st.container(border=True):
                    # Cabeçalho do painel
                    pa, pb = st.columns([6, 1])
                    with pa:
                        st.markdown(
                            f"<div class='p-title'>{sv(vm,'modelo')}"
                            f"<span class='p-fab'>{sv(vm,'fabricante')}</span></div>"
                            f"<div class='p-sub'>"
                            f"🔑 {sv(vm,'chassi')} &nbsp;·&nbsp; "
                            f"🪪 {sv(vm,'placa')} &nbsp;·&nbsp; "
                            f"🎨 {sv(vm,'cor')} &nbsp;·&nbsp; "
                            f"🏢 {sv(vm,'locadora')}"
                            f"</div>"
                            f"<span class='p-badge' style='background:{cor_p}'>{sv(vm,'status')}</span>"
                            f"&nbsp;<span style='font-size:11px;color:#94a3b8'>"
                            f"{farol_idade(id_p)} {f'{id_p}d' if id_p else '—'}</span>",
                            unsafe_allow_html=True,
                        )
                    with pb:
                        if st.button("✕ Fechar", key=f"px_{ch_r}", use_container_width=True):
                            st.session_state["gv_sel"] = None; st.rerun()

                    # Abas de ação — edição restrita a pode_editar
                    if pode_editar:
                        abas_p = ["🔄 Status", "📅 Agendar", "✏️ Editar", "📋 Detalhes", "🗑️ Deletar"]
                    else:
                        abas_p = ["📋 Detalhes"]
                    tp = st.tabs(abas_p)

                    # Índices dinâmicos
                    idx_status  = 0 if pode_editar else None
                    idx_agendar = 1 if pode_editar else None
                    idx_editar  = 2 if pode_editar else None
                    idx_detalhes= 3 if pode_editar else 0
                    idx_deletar = 4 if pode_editar else None

                    # ── Status ────────────────────────────
                    if pode_editar:
                      with tp[idx_status]:
                        with st.form(f"fp_s_{ch_r}"):
                            is_ = GV_STATUS_LIST.index(sv(vm,"status")) if sv(vm,"status") in GV_STATUS_LIST else 0
                            ns  = st.selectbox("Novo status", GV_STATUS_LIST, index=is_)
                            if st.form_submit_button("✅ Salvar status", use_container_width=True, type="primary"):
                                agr = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                                pg  = st.progress(0, "..."); pg.progress(50, "Salvando...")
                                gv_enviar({"aba":"veiculos","acao":"atualizar_linha","linha_num":lvm,"valores":[
                                    {"col":GV_COLUNAS.index("status")+1,         "valor":ns},
                                    {"col":GV_COLUNAS.index("atualizado_em")+1,  "valor":agr},
                                    {"col":GV_COLUNAS.index("atualizado_por")+1, "valor":st.session_state.get("auth_nome","Sistema")},
                                ]})
                                gv_enviar({"aba":"historico","acao":"inserir","linha":[
                                    sv(vm,"id"), sv(vm,"chassi"), sv(vm,"modelo"),
                                    sv(vm,"status"), ns, agr, st.session_state.get("auth_nome","Sistema"),
                                ]})
                                pg.progress(100, "✓"); gv_carregar.clear()
                                st.success(f"✅ Status → {ns}"); st.rerun()

                    # ── Agendar ──────────────────────────
                    if pode_editar:
                      with tp[idx_agendar]:
                        with st.form(f"fp_a_{ch_r}"):
                            a1, a2, a3 = st.columns(3)
                            with a1: nd  = st.date_input("Data *", value=hoje, format="DD/MM/YYYY")
                            with a2: nh  = st.time_input("Hora *", value=datetime.time(10,0))
                            with a3: nl_ = st.selectbox("Loja *", GV_LOJAS)
                            a4, a5 = st.columns(2)
                            with a4: ne = st.text_input("Entregador", value=sv(vm,"entregador") if sv(vm,"entregador")!="—" else "")
                            with a5: nc = st.text_input("Consultor",  value=sv(vm,"consultor")  if sv(vm,"consultor") !="—" else "")
                            conf = st.form_submit_button("📅 Confirmar agendamento", use_container_width=True, type="primary")

                        if conf:
                            errs = [e for e in [
                                verificar_rodizio(sv(vm,"placa"), nd, nh),
                                verificar_conflito_loja(df_gv, nd, nh, nl_, excluir_idx=idx_vm),
                            ] if e]
                            if errs:
                                for e in errs: st.error(e)
                                sf = st.text_input("Senha para forçar", type="password", key=f"ag_sf_{ch_r}")
                                if st.button("🔓 Forçar", key=f"ag_frc_{ch_r}"):
                                    if sf == SENHA_FECHAMENTO:
                                        st.session_state["ag_forcar"] = True; st.rerun()
                                    else: st.error("Senha incorreta.")
                            else:
                                st.session_state["ag_forcar"] = True

                        if st.session_state.get("ag_forcar"):
                            st.session_state["ag_forcar"] = False
                            agr = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                            pg2 = st.progress(0, "..."); pg2.progress(50, "Agendando...")
                            gv_enviar({"aba":"veiculos","acao":"atualizar_linha","linha_num":lvm,"valores":[
                                {"col":GV_COLUNAS.index("status")+1,         "valor":"Agendado"},
                                {"col":GV_COLUNAS.index("data_entrega")+1,   "valor":nd.strftime("%d/%m/%Y")},
                                {"col":GV_COLUNAS.index("hora_entrega")+1,   "valor":nh.strftime("%H:%M")},
                                {"col":GV_COLUNAS.index("loja_entrega")+1,   "valor":nl_},
                                {"col":GV_COLUNAS.index("entregador")+1,     "valor":ne},
                                {"col":GV_COLUNAS.index("consultor")+1,      "valor":nc},
                                {"col":GV_COLUNAS.index("atualizado_em")+1,  "valor":agr},
                                {"col":GV_COLUNAS.index("atualizado_por")+1, "valor":st.session_state.get("auth_nome","Sistema")},
                            ]})
                            gv_enviar({"aba":"historico","acao":"inserir","linha":[
                                sv(vm,"id"), sv(vm,"chassi"), sv(vm,"modelo"),
                                sv(vm,"status"), "Agendado", agr, st.session_state.get("auth_nome","Sistema"),
                            ]})
                            pg2.progress(100, "✓"); gv_carregar.clear()
                            st.success("✅ Agendado!"); st.rerun()

                    # ── Editar ────────────────────────────────────────────
                    if pode_editar:
                      with tp[idx_editar]:
                        with st.form(f"fp_e_{ch_r}"):
                            st.markdown("**🔍 Identificação**")
                            i1, i2, i3 = st.columns(3)
                            with i1:
                                ef  = st.selectbox("Fabricante", GV_FABRICANTES,
                                    index=GV_FABRICANTES.index(sv(vm,"fabricante")) if sv(vm,"fabricante") in GV_FABRICANTES else 0)
                                emd = st.text_input("Modelo", value=sv(vm,"modelo") if sv(vm,"modelo")!="—" else "")
                            with i2:
                                ep  = st.text_input("Placa", value=sv(vm,"placa") if sv(vm,"placa")!="—" else "")
                                ec  = st.text_input("Cor",   value=sv(vm,"cor")   if sv(vm,"cor")  !="—" else "")
                            with i3:
                                eaf = st.text_input("Ano Fab.", value=sv(vm,"ano_fabricacao") if sv(vm,"ano_fabricacao")!="—" else "")
                                eam = st.text_input("Ano Mod.", value=sv(vm,"ano_modelo")     if sv(vm,"ano_modelo")    !="—" else "")
                            ecb = st.selectbox("Combustível", GV_COMBUSTIVEIS,
                                index=GV_COMBUSTIVEIS.index(sv(vm,"combustivel")) if sv(vm,"combustivel") in GV_COMBUSTIVEIS else 0)
                            eopc = st.text_input("Opcionais", value=sv(vm,"opcionais") if sv(vm,"opcionais")!="—" else "")

                            st.markdown("**🏢 Operacional**")
                            o1, o2, o3 = st.columns(3)
                            with o1:
                                eloc = st.selectbox("Locadora", GV_LOCADORAS,
                                    index=GV_LOCADORAS.index(sv(vm,"locadora")) if sv(vm,"locadora") in GV_LOCADORAS else 0)
                                eco  = st.text_input("Consultor", value=sv(vm,"consultor") if sv(vm,"consultor")!="—" else "")
                            with o2:
                                ecl  = st.text_input("Cliente",     value=sv(vm,"cliente")     if sv(vm,"cliente")    !="—" else "")
                                epe  = st.text_input("Nº Pedido",   value=sv(vm,"pedido")      if sv(vm,"pedido")     !="—" else "")
                            with o3:
                                elo  = st.text_input("Local Atual", value=sv(vm,"local_atual") if sv(vm,"local_atual")!="—" else "")
                                elj  = st.selectbox("Loja Entrega", [""] + GV_LOJAS,
                                    index=([""] + GV_LOJAS).index(sv(vm,"loja_entrega")) if sv(vm,"loja_entrega") in GV_LOJAS else 0)
                            eav  = st.selectbox("Avaria?", ["Não","Sim"], index=1 if sv(vm,"avaria")=="Sim" else 0)
                            eob  = st.text_area("Obs. Avaria", value=sv(vm,"obs_avaria") if sv(vm,"obs_avaria")!="—" else "", height=68)

                            st.markdown("**📅 Datas e Entrega**")
                            d1, d2, d3 = st.columns(3)
                            with d1:
                                dch_v = parse_data(sv(vm,"data_chegada"))
                                ech   = st.date_input("Data Chegada", value=dch_v, format="DD/MM/YYYY")
                            with d2:
                                det_v = parse_data(sv(vm,"data_entrega"))
                                eet   = st.date_input("Data Entrega", value=det_v, format="DD/MM/YYYY")
                            with d3:
                                hora_str = sv(vm,"hora_entrega")
                                try:
                                    hh, mm = map(int, hora_str.split(":")); hv = datetime.time(hh, mm)
                                except:
                                    hv = datetime.time(10, 0)
                                eht = st.time_input("Hora Entrega", value=hv)
                            een = st.text_input("Entregador", value=sv(vm,"entregador") if sv(vm,"entregador")!="—" else "")

                            st.markdown("**💰 Financeiro**")
                            fn1, fn2, fn3 = st.columns(3)
                            with fn1:
                                try: nf_v = float(sv(vm,"valor_nf"))
                                except: nf_v = 0.0
                                enf = st.number_input("Valor NF (R$)", value=nf_v, min_value=0.0, step=100.0)
                            with fn2:
                                try: mg_v = float(sv(vm,"margem"))
                                except: mg_v = 0.0
                                emg = st.number_input("Margem (%)", value=mg_v, min_value=0.0, step=0.1)
                            with fn3:
                                try: cm_v = float(sv(vm,"comissao"))
                                except: cm_v = 0.0
                                ecm = st.number_input("Comissão (%)", value=cm_v, min_value=0.0, step=0.1)

                            if st.form_submit_button("💾 Salvar alterações", use_container_width=True, type="primary"):
                                agr = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                                pg3 = st.progress(0, "..."); pg3.progress(50, "Salvando...")
                                campos = [
                                    ("fabricante",  ef),
                                    ("modelo",      emd),
                                    ("placa",       ep),
                                    ("cor",         ec),
                                    ("ano_fabricacao", eaf),
                                    ("ano_modelo",  eam),
                                    ("combustivel", ecb),
                                    ("opcionais",   eopc),
                                    ("locadora",    eloc),
                                    ("consultor",   eco),
                                    ("cliente",     ecl),
                                    ("pedido",      epe),
                                    ("local_atual", elo),
                                    ("loja_entrega",elj),
                                    ("avaria",      eav),
                                    ("obs_avaria",  eob if eav=="Sim" else ""),
                                    ("data_chegada",ech.strftime("%d/%m/%Y") if ech else ""),
                                    ("data_entrega",eet.strftime("%d/%m/%Y") if eet else ""),
                                    ("hora_entrega",eht.strftime("%H:%M")),
                                    ("entregador",  een),
                                    ("valor_nf",    enf),
                                    ("margem",      emg),
                                    ("comissao",    ecm),
                                    ("atualizado_em",  agr),
                                    ("atualizado_por", st.session_state.get("auth_nome","Sistema")),
                                ]
                                vals = [{"col":GV_COLUNAS.index(c)+1,"valor":v} for c,v in campos if c in GV_COLUNAS]
                                ok   = gv_enviar({"aba":"veiculos","acao":"atualizar_linha","linha_num":lvm,"valores":vals})
                                pg3.progress(100, "✓")
                                if ok: gv_carregar.clear(); st.success("✅ Salvo!"); st.rerun()

                    # ── Detalhes ──────────────────────────
                    with tp[idx_detalhes]:
                        def vf(l, v):
                            cl = "em" if v == "—" else ""
                            return (
                                f"<div>"
                                f"<div class='det-l'>{l}</div>"
                                f"<div class='det-v {cl}'>{v}</div>"
                                f"</div>"
                            )
                        secs = [
                            ("🔍 Identificação", [
                                ("Fabricante", sv(vm,"fabricante")), ("Modelo", sv(vm,"modelo")),
                                ("Chassi", sv(vm,"chassi")),         ("Placa", sv(vm,"placa")),
                                ("Cor", sv(vm,"cor")),               ("Combustível", sv(vm,"combustivel")),
                                ("Ano Fab.", sv(vm,"ano_fabricacao")),("Ano Mod.", sv(vm,"ano_modelo")),
                                ("Opcionais", sv(vm,"opcionais")),
                            ]),
                            ("🏢 Operacional", [
                                ("Status", sv(vm,"status")),         ("Locadora", sv(vm,"locadora")),
                                ("Consultor", sv(vm,"consultor")),   ("Cliente", sv(vm,"cliente")),
                                ("Pedido", sv(vm,"pedido")),         ("Local Atual", sv(vm,"local_atual")),
                                ("Avaria", sv(vm,"avaria")),         ("Obs. Avaria", sv(vm,"obs_avaria")),
                            ]),
                            ("📅 Datas", [
                                ("Data Chegada", sv(vm,"data_chegada")),  ("Data Entrega", sv(vm,"data_entrega")),
                                ("Hora Entrega", sv(vm,"hora_entrega")),  ("Loja Entrega", sv(vm,"loja_entrega")),
                                ("Entregador", sv(vm,"entregador")),      ("Idade", f"{id_p}d" if id_p else "—"),
                            ]),
                            ("💰 Financeiro", [
                                ("Valor NF", sv(vm,"valor_nf")), ("Margem", sv(vm,"margem")), ("Comissão", sv(vm,"comissao")),
                            ]),
                            ("🕐 Auditoria", [
                                ("ID", sv(vm,"id")), ("Criado em", sv(vm,"criado_em")),
                                ("Atualizado em", sv(vm,"atualizado_em")), ("Por", sv(vm,"atualizado_por")),
                            ]),
                        ]
                        for tt, cc in secs:
                            st.markdown(
                                f"<div class='det-sec'>"
                                f"<div class='det-head'>{tt}</div>"
                                f"<div class='det-grid'>{''.join(vf(l,v) for l,v in cc)}</div>"
                                f"</div>",
                                unsafe_allow_html=True,
                            )

                    # ── Deletar ───────────────────────────
                    # ── Deletar ───────────────────────────
                    if pode_editar:
                      with tp[idx_deletar]:
                            st.error(
                                f"⚠️ Deletar permanentemente **{sv(vm,'modelo')}** ({sv(vm,'chassi')})?"
                            )
                            d1, d2 = st.columns(2)
                            with d1:
                                if st.button("✅ Sim, deletar", key=f"dp_{ch_r}",
                                             use_container_width=True, type="primary"):
                                    pg4 = st.progress(0, "..."); pg4.progress(60, "Removendo...")
                                    gv_enviar({"aba":"veiculos","acao":"deletar_linha","linha_num":lvm})
                                    gv_enviar({"aba":"historico","acao":"inserir","linha":[
                                        sv(vm,"id"), sv(vm,"chassi"), sv(vm,"modelo"),
                                        sv(vm,"status"), "DELETADO",
                                        datetime.datetime.now().strftime("%d/%m/%Y %H:%M"),
                                        st.session_state.get("auth_nome","Sistema"),
                                    ]})
                                    pg4.progress(100, "✓"); gv_carregar.clear()
                                    st.session_state["gv_sel"] = None; st.rerun()
                            with d2:
                                if st.button("❌ Cancelar", key=f"dc_{ch_r}", use_container_width=True):
                                    st.session_state["gv_sel"] = None; st.rerun()

    with aba_principal[1]:
        # ── Filtros da Agenda ────────────────────────────────
        st.markdown(f"<p style='font-size:13px;color:#94a3b8;margin-bottom:16px'>Agendamentos e entregas filtrados por data e grupo.</p>", unsafe_allow_html=True)

        ag1, ag2, ag3, ag4 = st.columns(4)
        with ag1:
            data_agenda = st.date_input("📅 Data", value=hoje, key="ag_data", format="DD/MM/YYYY")
        with ag2:
            opts_loc = ["Todas"] + sorted(df_gv["locadora"].dropna().unique().tolist()) if "locadora" in df_gv.columns else ["Todas"]
            fag_loc  = st.selectbox("Locadora", opts_loc, key="ag_loc")
        with ag3:
            opts_fab = ["Todas"] + sorted(df_gv["fabricante"].dropna().unique().tolist()) if "fabricante" in df_gv.columns else ["Todas"]
            fag_fab  = st.selectbox("Fabricante", opts_fab, key="ag_fab")
        with ag4:
            opts_loja = ["Todas"] + GV_LOJAS
            fag_loja  = st.selectbox("Loja", opts_loja, key="ag_loja")

        ag5, ag6 = st.columns(2)
        with ag5:
            opts_ent = ["Todos"] + sorted(df_gv["entregador"].dropna().unique().tolist()) if "entregador" in df_gv.columns else ["Todos"]
            fag_ent  = st.selectbox("Entregador", opts_ent, key="ag_ent")
        with ag6:
            opts_sta = ["Todos", "Agendado", "Reagendar", "Entregue"]
            fag_sta  = st.selectbox("Status", opts_sta, key="ag_sta")

        # Filtra
        data_str = data_agenda.strftime("%d/%m/%Y")
        df_ag = df_gv[df_gv["status"].isin(["Agendado","Reagendar","Entregue"])].copy()
        df_ag = df_ag[df_ag["data_entrega"].astype(str) == data_str]

        if fag_loc  != "Todas": df_ag = df_ag[df_ag["locadora"]    == fag_loc]
        if fag_fab  != "Todas": df_ag = df_ag[df_ag["fabricante"]  == fag_fab]
        if fag_loja != "Todas": df_ag = df_ag[df_ag["loja_entrega"]== fag_loja]
        if fag_ent  != "Todos": df_ag = df_ag[df_ag["entregador"]  == fag_ent]
        if fag_sta  != "Todos": df_ag = df_ag[df_ag["status"]      == fag_sta]

        if not df_ag.empty:
            df_ag["_f"] = df_ag.apply(farol_agendamento, axis=1)
            df_ag = df_ag.sort_values("hora_entrega", na_position="last")

        # KPIs da Agenda
        tot_ag   = len(df_ag) if not df_ag.empty else 0
        agend_ag = len(df_ag[df_ag["status"]=="Agendado"])   if not df_ag.empty else 0
        reag_ag  = len(df_ag[df_ag["status"]=="Reagendar"])  if not df_ag.empty else 0
        entr_ag  = len(df_ag[df_ag["status"]=="Entregue"])   if not df_ag.empty else 0
        atras_ag = len(df_ag[df_ag["_f"]=="🔴"])             if not df_ag.empty else 0

        st.markdown(f"""
        <div class="kpi-row">
          <div class="kpi-box"><div class="kpi-n" style="color:{AZUL}">{tot_ag}</div><div class="kpi-l">Total</div></div>
          <div class="kpi-box"><div class="kpi-n" style="color:#06b6d4">{agend_ag}</div><div class="kpi-l">Agendados</div></div>
          <div class="kpi-box"><div class="kpi-n" style="color:#f59e0b">{reag_ag}</div><div class="kpi-l">Reagendar</div></div>
          <div class="kpi-box"><div class="kpi-n" style="color:#10b981">{entr_ag}</div><div class="kpi-l">Entregues</div></div>
          <div class="kpi-box"><div class="kpi-n" style="color:#dc2626">{atras_ag}</div><div class="kpi-l">Atrasados</div></div>
        </div>
        """, unsafe_allow_html=True)

        st.divider()

        if df_ag.empty:
            st.info(f"Nenhum agendamento para {data_str} com os filtros selecionados.")
        else:
            for _, row in df_ag.iterrows():
                farol  = row.get("_f", "⚪")
                cor_st = STATUS_CORES.get(str(row.get("status","")), "#94a3b8")

                col_card, col_seg, col_pdf = st.columns([7, 3, 1])
                with col_card:
                    st.markdown(
                        f"<div style='background:#fff;border:1.5px solid #e8e0d0;"
                        f"border-left:5px solid {cor_st};"
                        f"border-radius:13px;padding:14px 20px;margin-bottom:8px;"
                        f"box-shadow:0 1px 5px rgba(0,0,0,.04)'>"
                        f"<div style='display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:12px'>"
                        f"<div style='flex:1;min-width:0'>"
                        f"<div style='font-size:15px;font-weight:700;color:{AZUL}'>"
                        f"{sv(row,'modelo')} <span style='font-size:12px;font-weight:400;color:#94a3b8'>{sv(row,'fabricante')}</span></div>"
                        f"<div style='display:flex;flex-wrap:wrap;gap:5px;margin-top:6px'>"
                        f"<span class='vtag'>🔑 {sv(row,'chassi')}</span>"
                        f"<span class='vtag'>🪪 {sv(row,'placa')}</span>"
                        f"<span class='vtag'>🎨 {sv(row,'cor')}</span>"
                        f"<span class='vtag'>🏢 {sv(row,'locadora')}</span>"
                        + (f"<span class='vtag'>📄 {sv(row,'pedido')}</span>" if sv(row,'pedido')!='—' else "") +
                        f"</div>"
                        f"<div style='font-size:11px;color:#475569;margin-top:5px'>"
                        f"👤 <b>{sv(row,'cliente')}</b> · {sv(row,'consultor')}</div>"
                        f"</div>"
                        f"<div style='text-align:right;flex-shrink:0'>"
                        f"<div style='font-size:26px;font-weight:800;color:{AZUL}'>{sv(row,'hora_entrega')}</div>"
                        f"<div style='font-size:11px;color:#64748b;margin-top:2px'>{sv(row,'loja_entrega')}</div>"
                        f"<div style='font-size:11px;color:#94a3b8'>🚚 {sv(row,'entregador')}</div>"
                        f"<span style='background:{cor_st};color:#fff;padding:3px 12px;"
                        f"border-radius:999px;font-size:11px;font-weight:700;"
                        f"display:inline-block;margin-top:6px'>{sv(row,'status')}</span>"
                        f"&nbsp;<span style='font-size:14px'>{farol}</span>"
                        f"</div></div></div>",
                        unsafe_allow_html=True
                    )
                with col_seg:
                    seg_key = f"seg_{sv(row,'chassi')}"
                    seg_sel = st.selectbox(
                        "📋 Segmento",
                        card_segmentos_disponiveis(),
                        key=seg_key,
                        help="Selecione o segmento para gerar o card correto"
                    )
                with col_pdf:
                    if st.button("📄", key=f"pdf_{sv(row,'chassi')}_{sv(row,'hora_entrega')}",
                                 use_container_width=True, help="Baixar card de agendamento"):
                        with st.spinner("Gerando card..."):
                            pdf_bytes = gerar_pdf_agendamento(row, sv, segmento=seg_sel)
                        mime_type = "application/pdf" if pdf_bytes[:4] == b'%PDF' else "text/plain"
                        ext       = "pdf" if mime_type == "application/pdf" else "txt"
                        st.download_button(
                            "⬇️", data=pdf_bytes,
                            file_name=f"card_{sv(row,'placa')}_{sv(row,'data_entrega').replace('/','')}.{ext}",
                            mime=mime_type,
                            use_container_width=True,
                            key=f"dl_{sv(row,'chassi')}_{sv(row,'hora_entrega')}",
                        )

    # ══════════════════════════════════════════════════════════
    # ABA RECEBIMENTO
    # ══════════════════════════════════════════════════════════
    if pode_editar:
     with aba_principal[2]:
        st.markdown(f"<h3 style='color:{AZUL};margin:0 0 4px'>📦 Recebimento de Veículos</h3>", unsafe_allow_html=True)
        st.markdown("<p style='color:#64748b;font-size:13px;margin-bottom:16px'>Cole a lista de chassis abaixo, um por linha. O sistema verifica na base e atualiza status e data de chegada.</p>", unsafe_allow_html=True)

        TRANSICAO = {
            "Trânsito Livre":   "Livre",
            "Trânsito Vendido": "Aguardando Agendamento",
        }

        lista_txt = st.text_area(
            "Chassis recebidos (um por linha)",
            placeholder="9BWBJ6BF5T4102813\n9BWBJ6BF1T4089378\n9BWBJ6BF2T4097375\n...",
            height=300,
            key="rec_lista"
        )

        chassis_validos = [c.strip().upper() for c in lista_txt.splitlines() if c.strip()]

        if chassis_validos:
            st.markdown(f"<p style='color:{AZUL};font-size:13px;font-weight:600'>{len(chassis_validos)} chassi(s) na lista</p>", unsafe_allow_html=True)

        if st.button("✅ Processar Recebimento", type="primary", key="rec_processar"):
            if not chassis_validos:
                st.error("Cole ao menos um chassi na lista.")
            else:
                hoje_str  = datetime.date.today().strftime("%d/%m/%Y")
                agora_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                usuario   = st.session_state.get("auth_nome", "Sistema")

                nao_encontrados = []
                status_errado   = []
                atualizados     = []

                prog = st.progress(0, "Processando...")
                for i, chassi in enumerate(chassis_validos):
                    prog.progress(int((i / len(chassis_validos)) * 90) + 5,
                                  f"Verificando {chassi}...")

                    if "chassi" not in df_gv.columns:
                        nao_encontrados.append(chassi); continue

                    match = df_gv[df_gv["chassi"].astype(str).str.upper() == chassi]
                    if match.empty:
                        nao_encontrados.append(chassi); continue

                    vm_rec     = match.iloc[0]
                    idx_rec    = match.index[0]
                    linha_rec  = int(idx_rec) + 2
                    status_atual = str(vm_rec.get("status","")).strip()

                    if status_atual not in TRANSICAO:
                        status_errado.append((chassi, status_atual, str(vm_rec.get("modelo","")))); continue

                    novo_status = TRANSICAO[status_atual]
                    ok = gv_enviar({"aba":"veiculos","acao":"atualizar_linha","linha_num":linha_rec,
                        "valores":[
                            {"col": GV_COLUNAS.index("status")+1,         "valor": novo_status},
                            {"col": GV_COLUNAS.index("data_chegada")+1,   "valor": hoje_str},
                            {"col": GV_COLUNAS.index("atualizado_em")+1,  "valor": agora_str},
                            {"col": GV_COLUNAS.index("atualizado_por")+1, "valor": usuario},
                        ]})
                    gv_enviar({"aba":"historico","acao":"inserir","linha":[
                        str(vm_rec.get("id","")), chassi,
                        str(vm_rec.get("modelo","")), status_atual,
                        novo_status, agora_str, usuario
                    ]})
                    if ok:
                        atualizados.append((chassi, str(vm_rec.get("modelo","")), status_atual, novo_status))

                prog.progress(100, "Concluído!")
                gv_carregar.clear()

                if atualizados:
                    st.success(f"✅ {len(atualizados)} veículo(s) recebido(s)!")
                    for ch, mod, de, para in atualizados:
                        st.markdown(
                            f"<div style='background:#f0fdf4;border:1px solid #bbf7d0;"
                            f"border-left:4px solid #22c55e;border-radius:8px;"
                            f"padding:8px 14px;margin-bottom:5px;font-size:13px'>"
                            f"<b style='color:#213144'>{mod}</b> &nbsp;·&nbsp;"
                            f"<span style='color:#64748b'>{ch}</span><br>"
                            f"<span style='font-size:11px;color:#64748b'>"
                            f"{de} → <b style='color:#22c55e'>{para}</b>"
                            f"&nbsp;·&nbsp; Chegada: {hoje_str}</span></div>",
                            unsafe_allow_html=True)

                if status_errado:
                    st.warning(f"⚠️ {len(status_errado)} chassi(s) com status incompatível:")
                    for ch, st_at, mod in status_errado:
                        st.markdown(
                            f"<div style='background:#fffbeb;border:1px solid #fde68a;"
                            f"border-left:4px solid #f59e0b;border-radius:8px;"
                            f"padding:8px 14px;margin-bottom:5px;font-size:13px'>"
                            f"<b style='color:#213144'>{mod}</b> &nbsp;·&nbsp;"
                            f"<span style='color:#64748b'>{ch}</span><br>"
                            f"<span style='font-size:11px;color:#92400e'>"
                            f"Status atual: <b>{st_at}</b> — sem transição automática</span></div>",
                            unsafe_allow_html=True)

                if nao_encontrados:
                    st.error(f"❌ {len(nao_encontrados)} chassi(s) não encontrado(s) — necessário cadastro:")
                    for ch in nao_encontrados:
                        st.markdown(
                            f"<div style='background:#fef2f2;border:1px solid #fecaca;"
                            f"border-left:4px solid #ef4444;border-radius:8px;"
                            f"padding:8px 14px;margin-bottom:5px;font-size:13px'>"
                            f"<b style='color:#ef4444'>{ch}</b>"
                            f"&nbsp;·&nbsp;<span style='font-size:11px;color:#64748b'>"
                            f"Não encontrado — cadastre via ➕ Cadastrar no Estoque</span></div>",
                            unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════
    # ABA E-MAIL DAS 11
    # ══════════════════════════════════════════════════════════
    if pode_editar:
     with aba_principal[3]:
        st.markdown(
            f"<h3 style='color:{AZUL};margin:0 0 4px'>📧 E-mail das 11</h3>"
            f"<p style='color:#64748b;font-size:13px;margin-bottom:16px'>"
            f"Solicitação de transporte — tabelas por loja prontas para copiar e colar.</p>",
            unsafe_allow_html=True
        )

        # ── Período ──────────────────────────────────────────
        dc1, dc2, _ = st.columns([2, 2, 5])
        with dc1: data_ini = st.date_input("📅 Data início", value=hoje, key="em_data_ini", format="DD/MM/YYYY")
        with dc2: data_fim = st.date_input("📅 Data fim",    value=hoje, key="em_data_fim", format="DD/MM/YYYY")

        if data_ini > data_fim:
            st.error("Data início não pode ser maior que data fim.")
            st.stop()

        datas_str_set = {
            (data_ini + datetime.timedelta(days=d)).strftime("%d/%m/%Y")
            for d in range((data_fim - data_ini).days + 1)
        }

        df_ag_email = df_gv[
            (df_gv["status"].isin(["Agendado", "Reagendar"])) &
            (df_gv["data_entrega"].astype(str).isin(datas_str_set))
        ].copy()
        if not df_ag_email.empty:
            df_ag_email = df_ag_email.sort_values(
                ["data_entrega","loja_entrega","hora_entrega"], na_position="last")

        # ── Extras ───────────────────────────────────────────
        st.markdown(
            f"<p style='font-size:13px;font-weight:600;color:{AZUL};margin:12px 0 4px'>"
            f"➕ Adicionar veículos não agendados</p>",
            unsafe_allow_html=True
        )
        n_extras = st.number_input("Quantidade", min_value=0, max_value=20,
                                    value=0, step=1, key="em_n_extras",
                                    label_visibility="collapsed")
        extras = []
        if n_extras > 0:
            for ei in range(int(n_extras)):
                ec1, ec2, ec3 = st.columns([3, 2, 2])
                with ec1: ch_extra = st.text_input(f"Chassi {ei+1}", placeholder="9BWBH6...", key=f"em_ch_{ei}")
                with ec2: lj_extra = st.selectbox("Loja", GV_LOJAS, key=f"em_lj_{ei}")
                with ec3: dt_extra = st.date_input("Data", value=hoje, key=f"em_dt_{ei}", format="DD/MM/YYYY")
                if ch_extra.strip():
                    ch_up = ch_extra.strip().upper()
                    match = df_gv[df_gv["chassi"].astype(str).str.upper() == ch_up]
                    if not match.empty:
                        r_ex = match.iloc[0].to_dict()
                        r_ex["loja_entrega"] = lj_extra
                        r_ex["data_entrega"] = dt_extra.strftime("%d/%m/%Y")
                        extras.append(r_ex)
                    else:
                        st.caption(f"⚠️ {ch_up} não encontrado na base.")

        st.divider()

        # ── Monta lista ───────────────────────────────────────
        linhas_todas = []
        for _, r in df_ag_email.iterrows():
            ja_solicitado = str(r.get("transporte_solicitado","")).strip() not in ("","—","nan","None")
            linhas_todas.append({
                "idx":     df_gv[df_gv["chassi"].astype(str)==sv(r,"chassi")].index[0]
                           if not df_gv[df_gv["chassi"].astype(str)==sv(r,"chassi")].empty else -1,
                "data":    sv(r,"data_entrega"),
                "loja":    sv(r,"loja_entrega"),
                "hora":    sv(r,"hora_entrega"),
                "cliente": sv(r,"cliente") if sv(r,"cliente") not in ("—","Estoque") else "—",
                "modelo":  sv(r,"modelo"),
                "chassi":  sv(r,"chassi"),
                "placa":   sv(r,"placa"),
                "cor":     sv(r,"cor"),
                "extra":   False,
                "ja_solicitado": ja_solicitado,
            })
        for r_ex in extras:
            def svx(c): return str(r_ex.get(c,"") or "").strip() or "—"
            linhas_todas.append({
                "idx":    -1,
                "data":   r_ex.get("data_entrega",""),
                "loja":   r_ex.get("loja_entrega",""),
                "hora":   "—",
                "cliente": svx("cliente") if svx("cliente") not in ("—","Estoque") else "—",
                "modelo":  svx("modelo"),
                "chassi":  svx("chassi"),
                "placa":   svx("placa"),
                "cor":     svx("cor"),
                "extra":   True,
                "ja_solicitado": False,
            })

        periodo_label = (data_ini.strftime("%d/%m/%Y") if data_ini == data_fim
                         else f"{data_ini.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}")

        if not linhas_todas:
            st.info(f"Nenhum agendamento para {periodo_label}.")
            texto_final = f"Solicitação de Transporte — {periodo_label}\n\nNenhuma entrega agendada."
        else:
            lojas_todas = sorted(set(l["loja"] for l in linhas_todas if l["loja"] not in ("","—")))
            total_veics = len(linhas_todas)
            ja_solicitados_base = sum(1 for l in linhas_todas if l["ja_solicitado"])

            st.markdown(
                f"<div style='background:#fdf8f0;border:1px solid #dfc28a;border-radius:10px;"
                f"padding:10px 16px;margin-bottom:12px;font-size:13px;color:{AZUL}'>"
                f"<b>{total_veics} veículo(s)</b> em <b>{len(lojas_todas)} loja(s)</b> · {periodo_label}"
                + (f" &nbsp;·&nbsp; <b style='color:#22c55e'>{ja_solicitados_base} já solicitados</b>" if ja_solicitados_base else "")
                + f"</div>",
                unsafe_allow_html=True
            )

            # ── Checklist por loja ────────────────────────────
            ck_key = "em_oks"
            if ck_key not in st.session_state:
                st.session_state[ck_key] = {}

            # Pré-marca os que já foram solicitados na base
            for l in linhas_todas:
                ck_id = f"ck_{l['chassi']}"
                if l["ja_solicitado"] and ck_id not in st.session_state[ck_key]:
                    st.session_state[ck_key][ck_id] = True

            chassis_marcados = set()

            for loja in lojas_todas:
                grupo = sorted([l for l in linhas_todas if l["loja"] == loja],
                               key=lambda x: (x["data"], x["hora"]))
                st.markdown(
                    f"<div style='font-size:13px;font-weight:700;color:{AZUL};"
                    f"border-bottom:2px solid {D_ESC};padding-bottom:4px;margin:14px 0 6px'>"
                    f"🏪 {loja} <span style='font-size:11px;font-weight:400;color:#64748b'>"
                    f"({len(grupo)} veículo(s))</span></div>",
                    unsafe_allow_html=True
                )

                # Tabela visual
                hds = ["✓","Data","Hora","Cliente","Modelo","Chassi","Placa","Cor"]
                tbl = (f"<table style='width:100%;border-collapse:collapse;font-size:12px;margin-bottom:6px'>"
                       f"<tr style='background:{AZUL};color:#fff'>"
                       + "".join(f"<th style='padding:6px 10px;text-align:left'>{h}</th>" for h in hds)
                       + "</tr>")
                for i, l in enumerate(grupo):
                    ck_id   = f"ck_{l['chassi']}"
                    checked = st.session_state[ck_key].get(ck_id, False)
                    if checked: chassis_marcados.add(l["chassi"])
                    bg      = "#f0fdf4" if checked else ("#fdf8f0" if l["extra"] else ("#f8f7f4" if i%2==0 else "#fff"))
                    icon    = "✅" if checked else ("🔔" if l["ja_solicitado"] else "⬜")
                    tag     = " <span style='font-size:10px;color:#b57b3f;font-weight:600'>[+]</span>" if l["extra"] else ""
                    tbl += (
                        f"<tr style='background:{bg}'>"
                        f"<td style='padding:5px 10px;text-align:center'>{icon}</td>"
                        f"<td style='padding:5px 10px;color:#64748b;white-space:nowrap'>{l['data']}</td>"
                        f"<td style='padding:5px 10px;color:#64748b'>{l['hora']}</td>"
                        f"<td style='padding:5px 10px'>{l['cliente']}{tag}</td>"
                        f"<td style='padding:5px 10px;font-weight:600;color:{AZUL}'>{l['modelo']}</td>"
                        f"<td style='padding:5px 10px;font-family:monospace;font-size:11px'>{l['chassi']}</td>"
                        f"<td style='padding:5px 10px'>{l['placa']}</td>"
                        f"<td style='padding:5px 10px'>{l['cor']}</td>"
                        f"</tr>"
                    )
                st.markdown(tbl + "</table>", unsafe_allow_html=True)

                # Checkboxes compactos
                n_cols = min(len(grupo), 3)
                ck_cols = st.columns(n_cols)
                for ci, l in enumerate(grupo):
                    ck_id = f"ck_{l['chassi']}"
                    with ck_cols[ci % n_cols]:
                        val = st.checkbox(
                            f"{l['chassi'][-7:]} · {l['placa']}",
                            value=st.session_state[ck_key].get(ck_id, False),
                            key=f"em_{ck_id}"
                        )
                        st.session_state[ck_key][ck_id] = val
                        if val: chassis_marcados.add(l["chassi"])

            # ── Botão confirmar solicitação na base ───────────
            st.divider()
            novos = [l for l in linhas_todas
                     if l["chassi"] in chassis_marcados and not l["ja_solicitado"] and l["idx"] >= 0]

            col_btn, col_info = st.columns([2, 5])
            with col_btn:
                confirmar = st.button(
                    f"✅ Confirmar solicitação ({len(novos)} novo(s))",
                    type="primary", key="em_confirmar",
                    disabled=len(novos) == 0
                )
            with col_info:
                if chassis_marcados:
                    st.markdown(
                        f"<p style='font-size:12px;color:#64748b;margin-top:8px'>"
                        f"<b>{len(chassis_marcados)}</b> marcado(s) · "
                        f"<b style='color:#22c55e'>{ja_solicitados_base}</b> já na base · "
                        f"<b style='color:#b57b3f'>{len(novos)}</b> serão gravados agora</p>",
                        unsafe_allow_html=True
                    )

            if confirmar and novos:
                agora_str = datetime.datetime.now().strftime("%d/%m/%Y %H:%M")
                usuario   = st.session_state.get("auth_nome","Sistema")
                pg = st.progress(0, "Gravando...")
                for i, l in enumerate(novos):
                    pg.progress(int((i/len(novos))*90)+5, f"Gravando {l['chassi']}...")
                    linha_num = int(l["idx"]) + 2
                    gv_enviar({"aba":"veiculos","acao":"atualizar_linha","linha_num":linha_num,
                        "valores":[
                            {"col": GV_COLUNAS.index("transporte_solicitado")+1,
                             "valor": agora_str},
                            {"col": GV_COLUNAS.index("local_atual")+1,
                             "valor": l["loja"]},
                            {"col": GV_COLUNAS.index("atualizado_em")+1,  "valor": agora_str},
                            {"col": GV_COLUNAS.index("atualizado_por")+1, "valor": usuario},
                        ]})
                pg.progress(100, "✓")
                gv_carregar.clear()
                st.success(f"✅ {len(novos)} chassis marcados como solicitados na base!")
                st.rerun()

            st.divider()

            # ── Texto para copiar ─────────────────────────────
            st.markdown(
                f"<p style='font-size:13px;font-weight:600;color:{AZUL};margin-bottom:6px'>"
                f"✉️ Texto pronto para copiar</p>",
                unsafe_allow_html=True
            )
            SEP  = "-" * 62
            SEP2 = "=" * 62
            linhas_txt = [
                f"SOLICITAÇÃO DE TRANSPORTE — {periodo_label}",
                SEP2,
                f"Total: {total_veics} veículo(s) | {len(lojas_todas)} loja(s)",
                "",
            ]
            for loja in lojas_todas:
                grupo = sorted([l for l in linhas_todas if l["loja"] == loja],
                               key=lambda x: (x["data"], x["hora"]))
                linhas_txt.append(f"📍 {loja.upper()} ({len(grupo)} veículo(s))")
                linhas_txt.append(SEP)
                linhas_txt.append(
                    f"{'OK':<4} {'DATA':<12} {'CLIENTE':<26} {'MODELO':<30} {'CHASSI':<20} {'PLACA':<10} {'COR'}")
                linhas_txt.append(SEP)
                for l in grupo:
                    ok_mark = "[✓]" if l["chassi"] in chassis_marcados else "[ ]"
                    extra   = " [+]" if l["extra"] else ""
                    linhas_txt.append(
                        f"{ok_mark:<4} {l['data']:<12} {(l['cliente']+extra)[:25]:<26} "
                        f"{l['modelo'][:29]:<30} {l['chassi'][:19]:<20} {l['placa'][:9]:<10} {l['cor']}"
                    )
                linhas_txt.append("")
            linhas_txt += [SEP2, "Carrera Signature"]
            texto_final = "\n".join(linhas_txt)

        st.text_area(
            "Selecione tudo (Ctrl+A) e copie (Ctrl+C)",
            value=texto_final, height=350, key="em_texto_final"
        )
